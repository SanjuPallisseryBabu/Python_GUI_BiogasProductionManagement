import control as ctrl
import datetime
import tkinter as tk
from tkinter import messagebox
import traceback
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import os
import matplotlib.dates as mdates
from scipy import stats
from sklearn.metrics import r2_score
from tkinter import ttk
from tkinter import filedialog  # For file dialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkcalendar import DateEntry
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from matplotlib.ticker import MaxNLocator
from matplotlib.dates import DateFormatter
import matplotlib.ticker as mticker
from scipy.ndimage import uniform_filter1d
from scipy.stats import zscore
from datetime import datetime, date, timedelta, time
from scipy import signal
import control
from matplotlib.ticker import AutoMinorLocator
from control import tf, c2d
from control.matlab import lsim
from datetime import timedelta
from scipy.interpolate import interp1d
from scipy.signal import butter, filtfilt



# Global variables for data segments
biogas_segment = None
substrate_segment = None
 
# Main window
root = tk.Tk()
root.title("Python GUI App")
root.geometry('1000x900')

# Set the window icon
icon_path = ''
root.iconbitmap(icon_path)

# Set background color of the main window
root.config(bg='#FFFFFF')

# Tabs
notebook = ttk.Notebook(root)
notebook.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

# Configure grid row and column weights to make the notebook expand
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# Create a ttk style object
style = ttk.Style()

# Use 'clam' theme to enable color customization
style.theme_use('clam')

# Configure style for the Notebook (tabs)
style.configure("TNotebook", background="#C7C7C7")  # Background color of the notebook
style.configure("TNotebook.Tab", 
                background="#C7C7C7",  # Tab background color
                foreground="black",      # Text color
                padding=[15, 4],         # Padding around text
                font=('Arial', 10),  # Font style
                borderwidth=0,           # Remove border width
                relief="flat"            # Flat relief for no border
               )


# Configure the active tab
style.map("TNotebook.Tab", 
          background=[("selected", "white")],  # Active tab background color
          foreground=[("selected", "black")], # Active tab text color
          borderwidth=[("selected", 0)],  # Set borderwidth to 0 for selected tab
           relief=[("selected", "flat")],  # Flat relief for no border
           padding=[("selected",[15, 5] )],
           font=[("selected", ('Arial', 10))]
         )

# Frames for Each Tab
load_tab = tk.Frame(notebook,bg="#F0F0F0")
preprocess_tab = tk.Frame(notebook)
model_tab = tk.Frame(notebook)
control_tab = tk.Frame(notebook)
feeding_tab = tk.Frame(notebook)

notebook.add(load_tab, text="Load Data")
notebook.add(preprocess_tab, text="Preprocessing")
notebook.add(model_tab, text="Model Estimation")
notebook.add(control_tab, text="Control System")
notebook.add(feeding_tab, text="Feeding Schedule")



#####
# Global DataFrames to store loaded data
gas_flow_rate_df = pd.DataFrame()
substrate_feeding_df = pd.DataFrame()

def load_data():
    global gas_flow_rate_df, substrate_feeding_df

    folder_path = "Data test Oktober 2023"
    if os.path.exists(folder_path):
        gas_flow_rate_files = []
        substrate_feeding_files = []

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file == "Gas flow rate Fermenter - current.csv":
                    gas_flow_rate_files.append(os.path.join(root, file))
                elif file == "Substrate feeding Fermenter - today.csv":
                    substrate_feeding_files.append(os.path.join(root, file))

        try:
            for file in gas_flow_rate_files:
                df = pd.read_csv(file, parse_dates=['TimeStamp'], dayfirst=True)
                gas_flow_rate_df = pd.concat([gas_flow_rate_df, df], ignore_index=True)
                print(f"Loaded Gas Flow Rate File: {file}")
                print(df.head())

            for file in substrate_feeding_files:
                df = pd.read_csv(file, parse_dates=['TimeStamp'], dayfirst=True)
                substrate_feeding_df = pd.concat([substrate_feeding_df, df], ignore_index=True)
                print(f"Loaded Substrate Feeding File: {file}")
                print(df.head())

            # Get both year and month from the CSV files
            first_timestamp = pd.to_datetime(gas_flow_rate_df['TimeStamp'].iloc[0])
            csv_year = first_timestamp.year
            csv_month = first_timestamp.month
            csv_day = first_timestamp.day

            # Create date object with CSV year and month, but current day
            csv_date = datetime(csv_year, csv_month, csv_day)
            
            # Update all DateEntry widgets with the CSV year and month
            start_date_down.set_date(csv_date)
            end_date_down.set_date(csv_date)
            start_date_up.set_date(csv_date)
            end_date_up.set_date(csv_date)

            radio_var.set("Yes")
            success_label.config(text="")
        except Exception as e:
            success_label.config(text=f"Error: {e}")
    else:
        success_label.config(text=f"Error: The folder '{folder_path}' does not exist.")

def plot_step_graph(upwards=False):
    global gas_flow_rate_df, substrate_feeding_df, fig1, fig2, canvas1, canvas2

    if gas_flow_rate_df.empty or substrate_feeding_df.empty:
        success_label.config(text="Error: Data not loaded. Please load the data first.")
        return

    try:
        # Get selected date range from calendar widgets for both plots
        if upwards:
            start_date = start_date_up.get_date()
            end_date = end_date_up.get_date()
            fig, canvas = fig2, canvas2  # Use second chart for upward plot
        else:
            start_date = start_date_down.get_date()
            end_date = end_date_down.get_date()
            fig, canvas = fig1, canvas1  # Use first chart for downward plot

        # Clear the current figure to prepare for new plot
        fig.clear()

        # Filter the data based on the selected date range
        mask_gas = (gas_flow_rate_df['TimeStamp'] >= pd.Timestamp(start_date)) & (gas_flow_rate_df['TimeStamp'] <= pd.Timestamp(end_date))
        mask_substrate = (substrate_feeding_df['TimeStamp'] >= pd.Timestamp(start_date)) & (substrate_feeding_df['TimeStamp'] <= pd.Timestamp(end_date))

        filtered_gas_flow_rate_df = gas_flow_rate_df.loc[mask_gas]
        filtered_substrate_feeding_df = substrate_feeding_df.loc[mask_substrate]

        # Plotting the graph with two y-axes
        ax1 = fig.add_subplot(111)

        # Plot Gas Flow Rate on the first y-axis with reduced thickness
        ax1.step(filtered_gas_flow_rate_df['TimeStamp'], filtered_gas_flow_rate_df['ValueNum'], where='post', 
                 label='Biogas Production Rate', linestyle='--', color='b', linewidth=0.8)  # Reduced thickness
        ax1.set_xlabel('Time', fontsize=12)
        ax1.set_ylabel('Biogas Production Rate [m3/h]', color='b', fontsize=12)
        ax1.tick_params(axis='y', labelcolor='b')

        # Customize x-axis to show date above time
        locator = mdates.HourLocator(byhour=[0, 12])  # Set ticks for 00:00 and 12:00
        formatter = mdates.DateFormatter('%d\n%H:%M')  # Multi-line label: '07\n00:00'
        ax1.xaxis.set_major_locator(locator)
        ax1.xaxis.set_major_formatter(formatter)

        # Add month and year to the last tick
        def add_month_year_to_last_tick(ax):
            ticks = ax.get_xticks()
            if len(ticks) > 0:
                ticks_as_dates = [mdates.num2date(tick) for tick in ticks]
                labels = [tick.strftime('%d\n%H:%M') for tick in ticks_as_dates]
                labels[-1] += f"\n{ticks_as_dates[-1].strftime('%b %Y')}"  # Append 'Oct 2023' to the last label
                ax.set_xticklabels(labels, ha='center')

        add_month_year_to_last_tick(ax1)

        # Ensure straight x-axis labels
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=0, ha='center')  # No rotation, labels are centered

        # Create a second y-axis for Substrate Feed Rate
        ax2 = ax1.twinx()
        
        # Interpolation for slanted transitions in Substrate Feeding
        time_substrate = filtered_substrate_feeding_df['TimeStamp'].values
        values_substrate = filtered_substrate_feeding_df['ValueNum'].values
        
        interpolated_time = []
        interpolated_values = []
        
        for i in range(len(time_substrate) - 1):
            # Add current point
            interpolated_time.append(time_substrate[i])
            interpolated_values.append(values_substrate[i])
            
            # Add interpolated slanting point
            midpoint_time = time_substrate[i] + (time_substrate[i + 1] - time_substrate[i]) / 2
            midpoint_value = (values_substrate[i] + values_substrate[i + 1]) / 2
            
            interpolated_time.append(midpoint_time)
            interpolated_values.append(midpoint_value)
        
        # Add the last point
        interpolated_time.append(time_substrate[-1])
        interpolated_values.append(values_substrate[-1])

        # Plot the interpolated Substrate Feeding line
        ax2.plot(interpolated_time, interpolated_values, label='Substrate Feeding', linestyle='-', color='r')
        ax2.set_ylabel('Substrate Feeding [t]', color='r', fontsize=12)
        ax2.tick_params(axis='y', labelcolor='r')

        # Adding legend
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines + lines2, labels + labels2, loc='upper left')

        # Adding title above the graph
        fig.suptitle("Before data preprocessing", fontsize=10, fontweight='bold')  # Bold title

        # Adjust layout to make room for the title
        fig.tight_layout(rect=[0, 0, 1, 0.95])  # Reserve space for the title
        ax1.grid(True)

        # Show the plot
        canvas.draw()
    except Exception as e:
        success_label.config(text=f"Error while plotting: {e}")



# CSV load
load_frame = tk.LabelFrame(load_tab,borderwidth=0, relief="flat",bg="#F0F0F0", padx=10, pady=0.5)
load_frame.grid(row=0, column=0, columnspan=2, padx=(10, 10), pady=0, sticky="nsew")

# Press button text
press_button = tk.Label(load_frame, text="Press the button to load CSV data", font=("Arial", 12, "bold"), background="#F0F0F0")
press_button.grid(row=0, column=0, padx=10, pady=(10,5), sticky="ew")

# Button
load_button = tk.Button(load_frame, text="Load Data", font=("Arial", 10), bd=1, relief="solid", command=load_data)
load_button.grid(row=1, column=0, padx=(20, 20), pady=3, ipadx=15, ipady=0) 

# Bind hover effects
#load_button.bind("<Enter>", on_enter)  # When mouse enters the button      
#load_button.bind("<Leave>", on_leave)  # When mouse leaves the button

# Add a label to display success or error messages
success_label = ttk.Label(load_tab, text="")
success_label.grid(pady=10)


# Load frame2 for radiobutton and labels
load_frame2 = tk.Frame(load_frame)
load_frame2.grid(row=0, column=1, rowspan=2, padx=0, pady=(10,5), sticky="nsew")

# Create variable for radio buttons
radio_var = tk.StringVar(value="No")  # Default to "No"

# Radio button for "No"
no_radio = tk.Radiobutton(load_frame2, text="No", variable=radio_var, value="No", bg="#F0F0F0", font=("Arial", 12))
no_radio.grid(row=0, column=0, padx=10, pady=(10,25), sticky="ns")

# Radio button for "Yes"
yes_radio = tk.Radiobutton(load_frame2, text="Yes", variable=radio_var, value="Yes", bg="#F0F0F0", font=("Arial", 12))
yes_radio.grid(row=0, column=2, padx=0, pady=(10,25), sticky="ns")

# Label to display the file selection status
status_label = tk.Label(root, text="")
status_label.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

# Move the "Data load is complete" label to column 1
toggle_label = tk.Label(load_frame, text="Data load is complete", font=("Arial", 10))
toggle_label.grid(row=1, column=1, padx=(10,40), pady=3, sticky="w")  

# Configure grid for load_frame
load_frame.grid_columnconfigure(0, weight=1)
load_frame.grid_columnconfigure(1, weight=1)
load_frame.grid_rowconfigure(0, weight=1)
load_frame.grid_rowconfigure(1, weight=1)

# Create a grid layout for the Step Downwards and Step Upwards sections
load_tab.grid_columnconfigure(0, weight=1)
load_tab.grid_columnconfigure(1, weight=1)

# Step Downwards section
down_frame = tk.LabelFrame(load_tab, borderwidth=0, relief="flat", bg="#F0F0F0",font=("Arial", 8))
down_frame.grid(row=2, column=0, padx=10, pady=(20,0), sticky='nsew')

# Center title within down_frame
title_label_down = tk.Label(down_frame, text="Select a time period for step downwards", font=("Arial", 10), bg="#F0F0F0", borderwidth=0, relief="flat")
title_label_down.grid(row=0, column=0, columnspan=4, padx=10, pady=10, sticky='nsew')

# Configure grid for down_frame
down_frame.grid_columnconfigure(0, weight=1)
down_frame.grid_columnconfigure(1, weight=1)
down_frame.grid_columnconfigure(2, weight=1)
down_frame.grid_columnconfigure(3, weight=1)

# Start and End date labels and DateEntry widgets in the same row for Step Down
start_label_down = tk.Label(down_frame, text="Start date", bg="#F0F0F0")
start_label_down.grid(row=1, column=0, padx=2, pady=5, sticky='e')

start_date_down = DateEntry(down_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
start_date_down.grid(row=1, column=1, padx=(2,0), pady=5, sticky='w')

end_label_down = tk.Label(down_frame, text="End date", bg="#F0F0F0")
end_label_down.grid(row=1, column=2, padx=(0,2), pady=5, sticky='e')

end_date_down = DateEntry(down_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
end_date_down.grid(row=1, column=3, padx=0, pady=5, sticky='w')

plot_down_button = tk.Button(down_frame, text="Plot data step downwards", font=("Arial", 10), bd=1, relief="solid",  command=lambda: plot_step_graph(upwards=False))
plot_down_button.grid(row=2, column=0, columnspan=4, padx=10, pady=(20,0), sticky='n')

def on_enter(event):
    plot_down_button['background'] = '#d0eaff'  # Light blue color on hover

def on_leave(event):
    plot_down_button['background'] = 'SystemButtonFace'  # Default button color
    
# Bind hover effects
plot_down_button.bind("<Enter>", on_enter)  # When mouse enters the button
plot_down_button.bind("<Leave>", on_leave)  # When mouse leaves the button

# Step Upwards section
up_frame = tk.LabelFrame(load_tab, borderwidth=0, relief="flat", bg="#F0F0F0", font=("Arial", 10))
up_frame.grid(row=2, column=1, padx=10, pady=(20, 20), sticky='nsew')  

# Center title within up_frame
title_label_up = tk.Label(up_frame, text="Select a time period for step upwards", font=("Arial", 10), bg="#F0F0F0", borderwidth=0, relief="flat")
title_label_up.grid(row=0, column=0, columnspan=4, padx=10, pady=10, sticky='nsew')

# Configure grid for up_frame
up_frame.grid_columnconfigure(0, weight=1)
up_frame.grid_columnconfigure(1, weight=1)
up_frame.grid_columnconfigure(2, weight=1)
up_frame.grid_columnconfigure(3, weight=1)

# Start and End date labels and DateEntry widgets in the same row for Step Up
start_label_up = tk.Label(up_frame, text="Start date", bg="#F0F0F0")
start_label_up.grid(row=1, column=0, padx=2, pady=5, sticky='e')

start_date_up = DateEntry(up_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
start_date_up.grid(row=1, column=1, padx=(2,0), pady=5, sticky='w')

end_label_up = tk.Label(up_frame, text="End date", bg="#F0F0F0")
end_label_up.grid(row=1, column=2, padx=(0,2), pady=5, sticky='e')

end_date_up = DateEntry(up_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
end_date_up.grid(row=1, column=3, padx=0, pady=5, sticky='w')

plot_up_button = tk.Button(up_frame, text="Plot data step upwards", font=("Arial", 10), bd=1, relief="solid", command=lambda: plot_step_graph(upwards=True))
plot_up_button.grid(row=2, column=0, columnspan=4, padx=10, pady=(20,0), sticky='n')

def on_enter(event):
    plot_up_button['background'] = '#d0eaff'  # Light blue color on hover

def on_leave(event):
    plot_up_button['background'] = 'SystemButtonFace'  # Default button color
    
# Bind hover effects
plot_up_button.bind("<Enter>", on_enter)  # When mouse enters the button
plot_up_button.bind("<Leave>", on_leave)  # When mouse leaves the button

# Label to display errors if any
error_label = ttk.Label(up_frame, text="", foreground="red")
error_label.grid()


# Create a new frame for charts
chart_frame = tk.LabelFrame(load_tab, borderwidth=0, relief="flat", bg="#F0F0F0", width=500, height=300)
chart_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=0, sticky="nsew")
chart_frame.grid_propagate(False)  # Prevent the frame from resizing to fit its contents
 
# Configure the chart frame for a 2-column layout
chart_frame.grid_columnconfigure(0, weight=1)
chart_frame.grid_columnconfigure(1, weight=1)
chart_frame.grid_rowconfigure(0, weight=1)

# Create empty figures for the charts
fig1, ax1 = plt.subplots()
fig2, ax2 = plt.subplots()

# Embed the first chart in the chart_frame
canvas1 = FigureCanvasTkAgg(fig1, master=chart_frame)
canvas1.get_tk_widget().grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
canvas1.draw()

# Embed the second chart in the chart_frame
canvas2 = FigureCanvasTkAgg(fig2, master=chart_frame)
canvas2.get_tk_widget().grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
canvas2.draw()

# Function to update the canvas size when the window is resized
def resize_canvas(event):
    fig1.set_size_inches(chart_frame.winfo_width() / 200, chart_frame.winfo_height() / 150)
    fig2.set_size_inches(chart_frame.winfo_width() / 200, chart_frame.winfo_height() / 150)
    canvas1.draw()
    canvas2.draw()

# Bind the resize event to update canvas size when the window is resized
chart_frame.bind("<Configure>", resize_canvas)

# Configure load_tab to expand the frame
load_tab.grid_rowconfigure(3, weight=1)
load_tab.grid_columnconfigure(0, weight=1)
load_tab.grid_columnconfigure(1, weight=1) 

# Create a frame with a grid layout inside load_tab
preprocess_frame = tk.LabelFrame(load_tab, borderwidth=0, relief="flat", bg="#F0F0F0")
preprocess_frame.grid(row=4, column=0,columnspan=4, padx=0, pady=0, sticky="nsew")

#PREPROCESSING TAB
##########################################################################################

# Preprocessing Control Variables
SMOOTHING_WINDOW = 5  # Window size for moving average
FILTER_ORDER = 4      # Butterworth filter order

def resample_and_interpolate(timestamps, values):
    """Resample and interpolate data to a uniform time series"""
    # Create uniform time grid
    time_min = timestamps.min()
    time_max = timestamps.max()
    uniform_timestamps = pd.date_range(start=time_min, end=time_max, freq='15T')  # Adjust frequency as needed
    
    # Interpolate values
    interpolator = interp1d(timestamps.astype(np.int64), values, kind='linear', fill_value='extrapolate')
    interpolated_values = interpolator(uniform_timestamps.astype(np.int64))
    
    return uniform_timestamps, interpolated_values

def calculate_precise_derivative(timestamps, values):
    """Calculate more precise first-order discrete derivative"""
    time_diffs = np.diff(timestamps.astype(np.int64)) / 1e9  # Convert to seconds
    value_diffs = np.diff(values)
    derivatives = value_diffs / time_diffs
    return timestamps[1:], derivatives

def design_butterworth_filter(data_timestamps):
    """Design Butterworth low-pass filter for biogas data"""
    # Calculate sampling frequency from timestamps
    time_diffs = np.diff(data_timestamps.astype(np.int64) // 10**9)
    fs = 1 / np.mean(time_diffs)  # Sampling frequency in Hz
    
    # Set cutoff frequency to target 2/3 hour periodic peaks
    cutoff_freq = 1/(2.5 * 3600)  # Convert to Hz
    
    # Design the filter
    nyquist = fs / 2
    normalized_cutoff = cutoff_freq / nyquist
    b, a = butter(FILTER_ORDER, normalized_cutoff, btype='low', analog=False)
    return b, a

def apply_preprocessing(data, timestamps):
    """Apply both low-pass filtering and moving average smoothing"""
    # Design and apply Butterworth filter
    b, a = design_butterworth_filter(timestamps)
    data_lowpass = filtfilt(b, a, data)
    
    # Apply moving average smoothing
    data_smoothed = uniform_filter1d(data_lowpass, size=SMOOTHING_WINDOW)
    return data_smoothed

# Function to change to the next tab
def next_tab():
    notebook.select(preprocess_tab)

# Function to dynamically update the layout for responsiveness
def make_responsive(widget, column_weights, row_weights):
    for col, weight in enumerate(column_weights):
        widget.grid_columnconfigure(col, weight=weight)
    for row, weight in enumerate(row_weights):
        widget.grid_rowconfigure(row, weight=weight)

# Responsive Layout for Preprocessing Tab
make_responsive(preprocess_tab, column_weights=[1, 1], row_weights=[0, 1])

# Global flags to track smoothing state
is_smoothed_down = False
is_smoothed_up = False

# Update the preprocessing button function
def preprocessing_button_pushed():
    global is_smoothed_down, is_smoothed_up
    is_smoothed_down = True
    is_smoothed_up = True
    update_step_downward_plot(smoothed=True)
    update_step_upward_plot(smoothed=True)
    next_tab()


# Configure grid for preprocess_frame to make it fully responsive
preprocess_frame.grid_columnconfigure(0, weight=1)
preprocess_frame.grid_columnconfigure(1, weight=0)  # Button in center
preprocess_frame.grid_columnconfigure(2, weight=1)

# Button for preprocessing in load tab
preprocess_button = tk.Button(preprocess_frame, text="Preprocessing", font=("Arial", 10), command=preprocessing_button_pushed)
preprocess_button.grid(row=0, column=1, padx=10, pady=10, ipadx=5, ipady=5, sticky="") 

### Preprocessing Tab with Two Columns ###
# Configure the grid layout for the preprocessing tab
preprocess_tab.grid_columnconfigure(0, weight=1)
preprocess_tab.grid_columnconfigure(1, weight=1)
preprocess_tab.grid_rowconfigure(2, weight=1)

# Add Titles to the First Row in Each Column
title_label_left = tk.Label(preprocess_tab, text="Preprocessed Data for Step Downwards", font=("Arial", 12, "bold"))
title_label_left.grid(row=0, column=0, padx=10, pady=10, sticky="n")

title_label_right = tk.Label(preprocess_tab, text="Preprocessed Data for Step Upwards", font=("Arial", 12, "bold"))
title_label_right.grid(row=0, column=1, padx=10, pady=10, sticky="n")

# Frame for Preprocessed Data for Step Downwards (Left Column)
step_down_frame = tk.Frame(preprocess_tab)
step_down_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
preprocess_tab.grid_rowconfigure(1, weight=1)

# Frame for Preprocessed Data for Step Upwards (Right Column)
step_up_frame = tk.Frame(preprocess_tab)
step_up_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
preprocess_tab.grid_rowconfigure(1, weight=1)

# Make frames responsive
make_responsive(step_down_frame, column_weights=[1], row_weights=[1, 0, 0])
make_responsive(step_up_frame, column_weights=[1], row_weights=[1, 0, 0])

# Matplotlib Figures for Step Downwards and Step Upwards
fig_down, ax_down = plt.subplots()
fig_up, ax_up = plt.subplots()

# Embed the Matplotlib figures in the Tkinter frames
canvas_down = FigureCanvasTkAgg(fig_down, master=step_down_frame)
canvas_down.get_tk_widget().grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

canvas_up = FigureCanvasTkAgg(fig_up, master=step_up_frame)
canvas_up.get_tk_widget().grid(row=0, column=0, padx=10, pady=10, sticky="nsew")


# Sliders for Step Downwards Section
slider_down_start = tk.Scale(step_down_frame, from_=0, to=100, orient="horizontal", label="Step Downwards Start")
slider_down_start.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

slider_down_end = tk.Scale(step_down_frame, from_=0, to=100, orient="horizontal", label="Step Downwards End")
slider_down_end.set(100)  # Set initial value to 100
slider_down_end.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

# Sliders for Step Upwards Section
slider_up_start = tk.Scale(step_up_frame, from_=0, to=100, orient="horizontal", label="Step Upwards Start")
slider_up_start.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

slider_up_end = tk.Scale(step_up_frame, from_=0, to=100, orient="horizontal", label="Step Upwards End")
slider_up_end.set(100)  # Set initial value to 100
slider_up_end.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

# Add the new slider change function
def on_slider_change(slider_type):
    """Handle slider value changes"""
    try:
        if slider_type == 'down':
            # Update preprocessing plot
            update_step_downward_plot()
            # Update model plot if data exists
            if global_vars['biogas_segment_down'] is not None:
                update_model_estimation_plots()
        else:  # slider_type == 'up'
            # Update preprocessing plot
            update_step_upward_plot()
            # Update model plot if data exists
            if global_vars['biogas_segment_up'] is not None:
                update_model_estimation_plots()
    except Exception as e:
        print(f"Error in slider change: {e}")

# Configure slider commands
slider_down_start.config(command=lambda val: on_slider_change('down'))
slider_down_end.config(command=lambda val: on_slider_change('down'))
slider_up_start.config(command=lambda val: on_slider_change('up'))
slider_up_end.config(command=lambda val: on_slider_change('up'))


# Configure the main preprocess_frame to center elements
preprocess_frame.grid_columnconfigure(0, weight=1)
preprocess_frame.grid_columnconfigure(1, weight=0)
preprocess_frame.grid_columnconfigure(2, weight=1)

global_vars = {
    'biogas_segment_down': None,
    'substrate_segment_down': None,
    'biogas_segment_up': None,
    'substrate_segment_up': None,
    'timestamps_down': None,
    'timestamps_up': None
}

def update_step_downward_plot(smoothed=False):
    """Update the step down plot with preprocessed data"""
    global is_smoothed_down
    if smoothed:
        is_smoothed_down = True

    try:
        # Parse start and end dates from user input
        start_date_str = start_date_down.get()
        end_date_str = end_date_down.get()
        start_date = datetime.strptime(start_date_str, "%m/%d/%y")
        end_date = datetime.strptime(end_date_str, "%m/%d/%y")

        # Filter data based on date range
        biogas_filtered = gas_flow_rate_df[
            (gas_flow_rate_df['TimeStamp'] >= start_date) &
            (gas_flow_rate_df['TimeStamp'] <= end_date)
        ]
        substrate_filtered = substrate_feeding_df[
            (substrate_feeding_df['TimeStamp'] >= start_date) &
            (substrate_feeding_df['TimeStamp'] <= end_date)
        ]

        # Store the full filtered data before applying sliders
        global_vars['full_biogas_down'] = biogas_filtered.copy()
        global_vars['full_substrate_down'] = substrate_filtered.copy()
        global_vars['full_timestamps_down'] = biogas_filtered['TimeStamp'].copy()

        # Apply slider values to refine the filtered data
        data_length = len(biogas_filtered)
        start_ind = round(slider_down_start.get() / 100 * data_length)
        end_ind = round(slider_down_end.get() / 100 * data_length)
        
        # Store slider indices for model estimation
        global_vars['slider_down_start_index'] = start_ind
        global_vars['slider_down_end_index'] = end_ind

        biogas_segment = biogas_filtered.iloc[start_ind:end_ind]
        substrate_segment = substrate_filtered.iloc[start_ind:end_ind]

        # NEW: Resample and Interpolate Data
        biogas_timestamps, biogas_values = resample_and_interpolate(
            biogas_segment['TimeStamp'], 
            biogas_segment['ValueNum']
        )
        substrate_timestamps, substrate_values = resample_and_interpolate(
            substrate_segment['TimeStamp'], 
            substrate_segment['ValueNum']
        )

        # Create new DataFrame with resampled data
        biogas_resampled = pd.DataFrame({
            'TimeStamp': biogas_timestamps,
            'ValueNum': biogas_values
        })
        substrate_resampled = pd.DataFrame({
            'TimeStamp': substrate_timestamps,
            'ValueNum': substrate_values
        })

        # Apply preprocessing if required
        value_col = 'ValueNum'
        if is_smoothed_down:
            # Calculate time differences for sampling frequency
            time_diffs = np.diff(biogas_segment['TimeStamp'].astype(np.int64) // 10**9)
            fs = 1 / np.mean(time_diffs)  # Sampling frequency in Hz
            
            # Design Butterworth filter
            cutoff_freq = 1/(2.5 * 3600)  # Target 2/3 hour periodic peaks
            nyquist = fs / 2
            normalized_cutoff = cutoff_freq / nyquist
            b, a = butter(4, normalized_cutoff, btype='low', analog=False)
            
            # Apply low-pass filter first
            biogas_data = biogas_segment[value_col].values
            substrate_data = substrate_segment[value_col].values
            
            # Handle any NaN values
            biogas_data = np.nan_to_num(biogas_data)
            substrate_data = np.nan_to_num(substrate_data)
            
            # Apply Butterworth filter
            biogas_lowpass = filtfilt(b, a, biogas_data)
            substrate_lowpass = filtfilt(b, a, substrate_data)
            
            # Then apply moving average smoothing
            window_size = 5  # Adjust window size as needed
            biogas_segment['SmoothedValueNum'] = uniform_filter1d(biogas_lowpass, size=window_size)
            substrate_segment['SmoothedValueNum'] = uniform_filter1d(substrate_lowpass, size=window_size)
            
            value_col = 'SmoothedValueNum'
            fig_down.suptitle("After data preprocessing", fontsize=10, fontweight='bold')
        else:
            fig_down.suptitle("Before data preprocessing", fontsize=10, fontweight='bold')

        # Store processed data in global variables for model estimation
        global_vars['biogas_segment_down'] = biogas_segment.copy()
        global_vars['timestamps_down'] = biogas_segment['TimeStamp'].copy()
        
        # Process substrate data
        substrate_segment['Derivative'] = substrate_segment['ValueNum'].diff().clip(lower=0)
        substrate_segment['FeedingRate'] = substrate_segment['Derivative'].fillna(0)
        
        # Handle sudden falls in FeedingRate
        threshold = 0.1
        for i in range(1, len(substrate_segment)):
            if substrate_segment['FeedingRate'].iloc[i] < threshold and substrate_segment['FeedingRate'].iloc[i - 1] > threshold:
                substrate_segment.loc[substrate_segment.index[i], 'FeedingRate'] = substrate_segment['FeedingRate'].iloc[i - 1]
        
        # Store processed substrate data
        global_vars['substrate_segment_down'] = substrate_segment.copy()

        # Clear the current figure to prepare for new plot
        fig_down.clear()
        ax1_down = fig_down.add_subplot(111)
        ax2_down = ax1_down.twinx()

        # Plot biogas production rate
        line1 = ax1_down.step(biogas_segment['TimeStamp'], 
                             biogas_segment[value_col], 
                             where='post', 
                             label='Biogas Production Rate',
                             linestyle='--',
                             color='b',
                             linewidth=0.8)

        # Configure axes
        ax1_down.set_xlabel('Time', fontsize=12)
        ax1_down.set_ylabel('Biogas Production Rate [m³/h]', color='b', fontsize=12)
        ax1_down.tick_params(axis='y', labelcolor='b')

        # Format time axis
        ax1_down.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%H:%M'))
        plt.setp(ax1_down.xaxis.get_majorticklabels(), rotation=0, ha='center')

        # Add month and year to last tick
        def add_month_year_to_last_tick(ax):
            ticks = ax.get_xticks()
            if len(ticks) > 0:
                ticks_as_dates = [mdates.num2date(tick) for tick in ticks]
                labels = [tick.strftime('%d\n%H:%M') for tick in ticks_as_dates]
                labels[-1] += f"\n{ticks_as_dates[-1].strftime('%b %Y')}"
                ax.set_xticklabels(labels, ha='center')

        add_month_year_to_last_tick(ax1_down)

        # Plot substrate feeding rate
        processed_substrate = substrate_segment[substrate_segment['FeedingRate'] > 0]
        line2 = ax2_down.plot(processed_substrate['TimeStamp'],
                             processed_substrate['FeedingRate'],
                             label='Substrate Feeding Rate',
                             color='r',
                             linewidth=1)
        
        ax2_down.set_ylabel('Substrate Feeding Rate [t/h]', color='r', fontsize=12)
        ax2_down.tick_params(axis='y', labelcolor='r')

        # Add legend
        lines = line1 + line2
        labels = [l.get_label() for l in lines]
        ax1_down.legend(lines, labels, loc='upper left')

        # Add grid
        ax1_down.grid(True)
        
        # Adjust layout and draw
        fig_down.tight_layout(rect=[0, 0, 1, 0.95])
        canvas_down.draw()

        # Update model estimation plots if needed
        if is_smoothed_down:
            update_model_estimation_plots()

    except Exception as e:
        print(f"Error during graph preprocessing and plotting: {e}")
        traceback.print_exc()  # Print full error traceback for debugging

def update_step_upward_plot(smoothed=False):
    """Update the step up plot with preprocessed data"""
    global is_smoothed_up
    if smoothed:
        is_smoothed_up = True

    try:
        # Parse start and end dates from user input
        start_date_str = start_date_up.get()
        end_date_str = end_date_up.get()
        start_date = datetime.strptime(start_date_str, "%m/%d/%y")
        end_date = datetime.strptime(end_date_str, "%m/%d/%y")

        # Filter data based on date range
        biogas_filtered = gas_flow_rate_df[
            (gas_flow_rate_df['TimeStamp'] >= start_date) &
            (gas_flow_rate_df['TimeStamp'] <= end_date)
        ]
        substrate_filtered = substrate_feeding_df[
            (substrate_feeding_df['TimeStamp'] >= start_date) &
            (substrate_feeding_df['TimeStamp'] <= end_date)
        ]

        # Store the full filtered data before applying sliders
        global_vars['full_biogas_up'] = biogas_filtered.copy()
        global_vars['full_substrate_up'] = substrate_filtered.copy()
        global_vars['full_timestamps_up'] = biogas_filtered['TimeStamp'].copy()

        # Apply slider values to refine the filtered data
        data_length = len(biogas_filtered)
        start_ind = round(slider_up_start.get() / 100 * data_length)
        end_ind = round(slider_up_end.get() / 100 * data_length)
        
        # Store slider indices for model estimation
        global_vars['slider_up_start_index'] = start_ind
        global_vars['slider_up_end_index'] = end_ind

        biogas_segment = biogas_filtered.iloc[start_ind:end_ind]
        substrate_segment = substrate_filtered.iloc[start_ind:end_ind]

        # NEW: Resample and Interpolate Data
        biogas_timestamps, biogas_values = resample_and_interpolate(
            biogas_segment['TimeStamp'], 
            biogas_segment['ValueNum']
        )
        substrate_timestamps, substrate_values = resample_and_interpolate(
            substrate_segment['TimeStamp'], 
            substrate_segment['ValueNum']
        )

        # Create new DataFrame with resampled data
        biogas_resampled = pd.DataFrame({
            'TimeStamp': biogas_timestamps,
            'ValueNum': biogas_values
        })
        substrate_resampled = pd.DataFrame({
            'TimeStamp': substrate_timestamps,
            'ValueNum': substrate_values
        })


        # Apply preprocessing if required
        value_col = 'ValueNum'
        if is_smoothed_up:
            # Calculate time differences for sampling frequency
            time_diffs = np.diff(biogas_segment['TimeStamp'].astype(np.int64) // 10**9)
            fs = 1 / np.mean(time_diffs)  # Sampling frequency in Hz
            
            # Design Butterworth filter
            cutoff_freq = 1/(2.5 * 3600)  # Target 2/3 hour periodic peaks
            nyquist = fs / 2
            normalized_cutoff = cutoff_freq / nyquist
            b, a = butter(4, normalized_cutoff, btype='low', analog=False)
            
            # Apply low-pass filter
            biogas_lowpass = filtfilt(b, a, biogas_segment[value_col].values)
            substrate_lowpass = filtfilt(b, a, substrate_segment[value_col].values)
            
            # Apply moving average smoothing
            window_size = 5
            biogas_segment['SmoothedValueNum'] = uniform_filter1d(biogas_lowpass, size=window_size)
            substrate_segment['SmoothedValueNum'] = uniform_filter1d(substrate_lowpass, size=window_size)
            value_col = 'SmoothedValueNum'
            
            # Update plot title
            fig_up.suptitle("After data preprocessing", fontsize=10, fontweight='bold')
        else:
            fig_up.suptitle("Before data preprocessing", fontsize=10, fontweight='bold')

        # Store processed data in global variables for model estimation
        global_vars['biogas_segment_up'] = biogas_segment.copy()
        global_vars['timestamps_up'] = biogas_segment['TimeStamp'].copy()
        
        # Process substrate data
        substrate_segment['Derivative'] = substrate_segment['ValueNum'].diff().clip(lower=0)
        substrate_segment['FeedingRate'] = substrate_segment['Derivative'].fillna(0)
        
        # Handle sudden falls in FeedingRate
        threshold = 0.1
        for i in range(1, len(substrate_segment)):
            if substrate_segment['FeedingRate'].iloc[i] < threshold and substrate_segment['FeedingRate'].iloc[i - 1] > threshold:
                substrate_segment.loc[substrate_segment.index[i], 'FeedingRate'] = substrate_segment['FeedingRate'].iloc[i - 1]
        
        # Store processed substrate data
        global_vars['substrate_segment_up'] = substrate_segment.copy()

        # Clear the current figure to prepare for new plot
        fig_up.clear()
        ax1_up = fig_up.add_subplot(111)
        ax2_up = ax1_up.twinx()

        # Plot biogas production rate
        line1 = ax1_up.step(biogas_segment['TimeStamp'], 
                           biogas_segment[value_col], 
                           where='post', 
                           label='Biogas Production Rate',
                           linestyle='--',
                           color='b',
                           linewidth=0.8)

        # Configure axes
        ax1_up.set_xlabel('Time', fontsize=12)
        ax1_up.set_ylabel('Biogas Production Rate [m³/h]', color='b', fontsize=12)
        ax1_up.tick_params(axis='y', labelcolor='b')

        # Format time axis
        ax1_up.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%H:%M'))
        plt.setp(ax1_up.xaxis.get_majorticklabels(), rotation=0, ha='center')

        # Add month and year to last tick
        def add_month_year_to_last_tick(ax):
            ticks = ax.get_xticks()
            if len(ticks) > 0:
                ticks_as_dates = [mdates.num2date(tick) for tick in ticks]
                labels = [tick.strftime('%d\n%H:%M') for tick in ticks_as_dates]
                labels[-1] += f"\n{ticks_as_dates[-1].strftime('%b %Y')}"
                ax.set_xticklabels(labels, ha='center')

        add_month_year_to_last_tick(ax1_up)

        # Plot substrate feeding rate
        processed_substrate = substrate_segment[substrate_segment['FeedingRate'] > 0]
        line2 = ax2_up.plot(processed_substrate['TimeStamp'],
                           processed_substrate['FeedingRate'],
                           label='Substrate Feeding Rate',
                           color='r',
                           linewidth=1)
        
        ax2_up.set_ylabel('Substrate Feeding Rate [t/h]', color='r', fontsize=12)
        ax2_up.tick_params(axis='y', labelcolor='r')

        # Add legend
        lines = line1 + line2
        labels = [l.get_label() for l in lines]
        ax1_up.legend(lines, labels, loc='upper left')

        # Add grid
        ax1_up.grid(True)
        
        # Adjust layout and draw
        fig_up.tight_layout(rect=[0, 0, 1, 0.95])
        canvas_up.draw()

        # Update model estimation plots if needed
        update_model_estimation_plots()

    except Exception as e:
        print(f"Error during graph preprocessing and plotting: {e}")
        traceback.print_exc()  # Print full error traceback for debugging

###############################################################################################################
#MODEL ESTIMATION

# Model Estimation Button in Preprocessing Tab
model_estimation_frame = tk.Frame(preprocess_tab)
model_estimation_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

# Add this to ensure data is preprocessed before model estimation
def model_estimation_button_pushed():
    """Handle transition to model estimation tab and initialize analysis"""
    if (global_vars['biogas_segment_down'] is None or 
        global_vars['substrate_segment_down'] is None or
        global_vars['biogas_segment_up'] is None or
        global_vars['substrate_segment_up'] is None):
        print("Please preprocess the data first before proceeding to model estimation.")
        return
    
    notebook.select(model_tab)  # Switch to model estimation tab
    update_model_estimation_plots()

model_estimation_button = tk.Button(
    model_estimation_frame, 
    text="Model Estimation",
    font=("Arial", 10),
    command=model_estimation_button_pushed
)
model_estimation_button.grid(row=0, column=0, padx=10, pady=10, ipadx=10, ipady=5, sticky="")

# Make the frame fully responsive and centered
preprocess_tab.grid_rowconfigure(3, weight=1)  # Add weight to the row where the button resides
model_estimation_frame.grid_columnconfigure(0, weight=1)  # Center horizontally
model_estimation_frame.grid_rowconfigure(0, weight=1)  # Center vertically

###################
# Configure the grid layout of the model estimation tab
model_tab.grid_columnconfigure(0, weight=1)
model_tab.grid_columnconfigure(1, weight=1)
model_tab.grid_rowconfigure(1, weight=1)
model_tab.grid_rowconfigure(2, weight=0)  # Button row

# Titles for Model Estimation Tab
title_label_down_model = tk.Label(
    model_tab, 
    text="Preprocessed data for step downwards",
    font=("Arial", 12, "bold")
)
title_label_down_model.grid(row=0, column=0, padx=10, pady=10, sticky="n")

title_label_up_model = tk.Label(
    model_tab, 
    text="Preprocessed data for step upwards",
    font=("Arial", 12, "bold")
)
title_label_up_model.grid(row=0, column=1, padx=10, pady=10, sticky="n")

# Frames for Charts and Tables
down_model_frame = tk.Frame(model_tab)
down_model_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")

up_model_frame = tk.Frame(model_tab)
up_model_frame.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")

# Make frames responsive
down_model_frame.grid_columnconfigure(0, weight=1)
down_model_frame.grid_rowconfigure(0, weight=2)
down_model_frame.grid_rowconfigure(1, weight=1)

up_model_frame.grid_columnconfigure(0, weight=1)
up_model_frame.grid_rowconfigure(0, weight=2)
up_model_frame.grid_rowconfigure(1, weight=1)

# Create Matplotlib figures for model estimation
fig_model_down, ax_model_down = plt.subplots(figsize=(6, 4))
fig_model_up, ax_model_up = plt.subplots(figsize=(6, 4))

# Embed figures in the frames
canvas_model_down = FigureCanvasTkAgg(fig_model_down, master=down_model_frame)
canvas_model_down.get_tk_widget().grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

canvas_model_up = FigureCanvasTkAgg(fig_model_up, master=up_model_frame)
canvas_model_up.get_tk_widget().grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# Add a title for correlation metrics for the downward step
correlation_title_down = tk.Label(
    down_model_frame,
    text="Correlation between selected model and preprocessed gas production flow rate",
    font=("Arial", 11, "bold"),
    anchor="center"
)
correlation_title_down.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

# Add a title for correlation metrics for the upward step
correlation_title_up = tk.Label(
    up_model_frame,
    text="Correlation between selected model and preprocessed gas production flow rate",
    font=("Arial", 11, "bold"),
    anchor="center"
)
correlation_title_up.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

# Make the rows for the titles in the frames responsive
down_model_frame.grid_rowconfigure(2, weight=1)
up_model_frame.grid_rowconfigure(2, weight=1)

# Make the columns for the titles in the frames responsive
down_model_frame.grid_columnconfigure(0, weight=1)
up_model_frame.grid_columnconfigure(0, weight=1)


# Create frames for metrics tables
metrics_frame_down = ttk.Frame(down_model_frame)
metrics_frame_down.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

metrics_frame_up = ttk.Frame(up_model_frame)
metrics_frame_up.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")


# Ensure the frames are fully responsive
down_model_frame.grid_rowconfigure(1, weight=1)
up_model_frame.grid_rowconfigure(1, weight=1)
metrics_frame_down.grid_columnconfigure(0, weight=1)
metrics_frame_up.grid_columnconfigure(0, weight=1)

# Create tables for similarity metrics
columns = ('Similarity metrics', 'Value')
tree_down = ttk.Treeview(metrics_frame_down, columns=columns, show='headings', height=7)
tree_up = ttk.Treeview(metrics_frame_up, columns=columns, show='headings', height=7)

for tree in [tree_down, tree_up]:
    tree.heading('Similarity metrics', text='Similarity metrics')
    tree.heading('Value', text='Value')
    tree.column('Similarity metrics', width=250)
    tree.column('Value', width=150)
    tree.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

# Add Select Model button at the bottom
select_model_frame = tk.Frame(model_tab)
select_model_frame.grid(row=2, column=0, columnspan=2, pady=10)

# Create a variable to store the selected model
selected_model = tk.StringVar(value="model1")  # Default selection

# Create a frame for radio buttons
radio_frame = tk.Frame(select_model_frame)
radio_frame.grid(row=0, column=0, columnspan=2, pady=5)

# Create and arrange radio buttons horizontally
models = [
    ("Time percentage", "model1"),
    ("Pt1_Modell", "model2"),
    ("Time constant sum", "model3"),
    ("Turning tangent", "model4"),
    ("Pt1 estimator", "model5"),
    ("Pt2 estimator", "model6")
]

for i, (text, value) in enumerate(models):
    rb = tk.Radiobutton(
        radio_frame,
        text=text,
        variable=selected_model,
        value=value,
        font=("Arial", 10)
    )
    rb.grid(row=0, column=i, padx=10)

# Ensure the Treeview fills the frame
metrics_frame_down.grid_rowconfigure(0, weight=1)
metrics_frame_down.grid_columnconfigure(0, weight=1)
metrics_frame_up.grid_rowconfigure(0, weight=1)
metrics_frame_up.grid_columnconfigure(0, weight=1)


# Configure the frame to center the button
select_model_frame.grid_columnconfigure(0, weight=1)  # Make column 0 responsive
select_model_frame.grid_rowconfigure(1, weight=1)    # Adjust the row height for centering

# Create and place the button with no extra width
select_model_button = tk.Button(
    select_model_frame,
    text="Select model",
    font=("Arial", 10),
    command=lambda: select_model(),
    anchor="center"  # Center the text within the button
)
select_model_button.grid(row=1, column=0, padx=10, pady=5)  # No sticky needed to avoid resizing

# Add hover effects for buttons
def on_enter_model(event):
    event.widget['background'] = '#d0eaff'  # Light blue color on hover

def on_leave_model(event):
    event.widget['background'] = 'SystemButtonFace'  # Default button color

# Bind hover effects to buttons
select_model_button.bind("<Enter>", on_enter_model)
select_model_button.bind("<Leave>", on_leave_model)

def calculate_metrics(measured, modeled):

    import numpy as np
    from scipy import stats
    
    try:
        metrics = {}
        
        # Handle NaN and infinite values
        valid_indices = ~(np.isnan(measured) | np.isnan(modeled) | 
                         np.isinf(measured) | np.isinf(modeled))
        measured_clean = np.array(measured)[valid_indices]
        modeled_clean = np.array(modeled)[valid_indices]
        
        # Make sure we have valid data
        if len(measured_clean) < 2 or len(modeled_clean) < 2:
            print("Warning: Not enough valid data points for metrics calculation")
            return {
                'R2': 0, 'Pearson': 0, 'Spearman': 0, 'Kendall': 0,
                'Euclidean': 0, 'Normalized_Euclidean': 0, 
                'Chebyshev': 0, 'Cosine': 0
            }
        
        # R² calculation using correlation coefficient squared
        correlation = np.corrcoef(measured_clean, modeled_clean)[0, 1]
        metrics['R2'] = correlation ** 2
        
        # Correlation coefficients
        metrics['Pearson'] = correlation
        metrics['Spearman'] = stats.spearmanr(measured_clean, modeled_clean)[0]
        metrics['Kendall'] = stats.kendalltau(measured_clean, modeled_clean)[0]
        
        # Calculate Euclidean distance over whole time series
        euclidean_distance = np.sqrt(np.sum((measured_clean - modeled_clean) ** 2))
        metrics['Euclidean'] = euclidean_distance
        
        # Calculate normalized Euclidean distance (relative to data range)
        data_range = np.max(measured_clean) - np.min(measured_clean)
        if data_range > 0:
            metrics['Normalized_Euclidean'] = euclidean_distance / (data_range * np.sqrt(len(measured_clean)))
        else:
            metrics['Normalized_Euclidean'] = 0
            
        # Calculate RMSE (Root Mean Square Error)
        metrics['RMSE'] = np.sqrt(np.mean((measured_clean - modeled_clean) ** 2))
 
        # Calculate Chebyshev distance as maximum single-point difference
        metrics['Chebyshev'] = np.max(np.abs(measured_clean - modeled_clean))
        
        # Cosine distance
        norm_measured = np.linalg.norm(measured_clean)
        norm_modeled = np.linalg.norm(modeled_clean)
        if norm_measured > 0 and norm_modeled > 0:
            metrics['Cosine'] = 1 - np.dot(measured_clean, modeled_clean) / (norm_measured * norm_modeled)
        else:
            metrics['Cosine'] = 0
            
        # Calculate MAE (Mean Absolute Error)
        metrics['MAE'] = np.mean(np.abs(measured_clean - modeled_clean))
        
        return metrics
        
    except Exception as e:
        print(f"Error in calculate_metrics: {e}")
        import traceback
        traceback.print_exc()
        return {
            'R2': 0,
            'Pearson': 0,
            'Spearman': 0, 
            'Kendall': 0,
            'Euclidean': 0,
            'Normalized_Euclidean': 0,
            'RMSE': 0,
            'MAE': 0,
            'Chebyshev': 0,
            'Cosine': 0
        }
    
#AIC Calculation for result
def calculate_aic_metrics(measured, modeled):
    """
    Calculate AIC and SSE for model evaluation
    """
    import numpy as np
    
    try:
        # Convert inputs to numpy arrays
        measured = np.array(measured)
        modeled = np.array(modeled)
        
        # Handle NaN and infinite values
        valid_indices = ~(np.isnan(measured) | np.isnan(modeled) | 
                         np.isinf(measured) | np.isinf(modeled))
        measured_clean = measured[valid_indices]
        modeled_clean = modeled[valid_indices]
        
        # Get number of data points
        n = len(measured_clean)
        
        if n < 4:  # Need at least a few points for meaningful calculation
            print("Warning: Not enough valid data points for AIC calculation")
            return {
                'AIC': float('inf'),
                'SSE': float('inf'),
                'n': n,
                'P': 3  # Default parameters for time percentage method
            }
        
        # Calculate SSE (Sum of Squares Error)
        residuals = measured_clean - modeled_clean
        SSE = np.sum(residuals**2)
        
        # Number of parameters in time percentage model
        # For the time percentage method: K (gain), TA, TB (time constants)
        P = 3
        
        # Calculate AIC
        # AIC = n * ln(SSE/n) + 2P
        AIC = 2 * np.log(SSE) + 2*P
        
        # Calculate corrected AIC for small sample sizes
        # AICc = AIC + (2P(P+1))/(n-P-1)
        AICc = AIC + (2*P*(P+1))/(n-P-1) if n > P+1 else float('inf')
        
        # Calculate MSE (Mean Square Error)
        MSE = SSE / n
        
        # Calculate RMSE (Root Mean Square Error)
        RMSE = np.sqrt(MSE)
        
        return {
            'AIC': AIC,
            'AICc': AICc,  # Corrected AIC for small sample sizes
            'SSE': SSE,
            'MSE': MSE,
            'RMSE': RMSE,
            'n': n,      # Number of data points
            'P': P       # Number of parameters
        }
        
    except Exception as e:
        print(f"Error in calculate_aic_metrics: {e}")
        import traceback
        traceback.print_exc()
        return {
            'AIC': float('inf'),
            'AICc': float('inf'),
            'SSE': float('inf'),
            'MSE': float('inf'),
            'RMSE': float('inf'),
            'n': 0,
            'P': 3
        }
        
def update_metrics_table(tree, metrics):
    """Update the metrics display in the treeview"""
    # Clear existing items
    for item in tree.get_children():
        tree.delete(item)
        
    # Add new metrics
    metrics_display = [
        ('Coefficient of Determination (R²)', f"{metrics['R2']:.4f}"),
        ('Pearson Correlation Coefficient', f"{metrics['Pearson']:.4f}"),
        ('Spearman Rank Correlation', f"{metrics['Spearman']:.4f}"),
        ('Kendall Rank Correlation', f"{metrics['Kendall']:.4f}"),
        ('Euclidean Distance', f"{metrics['Euclidean']:.4f}"),
        ('Chebyshev Distance', f"{metrics['Chebyshev']:.4f}"),
        ('Cosine Distance', f"{metrics['Cosine']:.4f}")
    ]
    
    for metric, value in metrics_display:
        tree.insert('', tk.END, values=(metric, value))


def calculate_zeitprozentkennwert(flowrate, feed, ind_sprung, feed_max, feed_min, flowrate_max, flowrate_min, timestamps, step_direction):
    """Calculate zeitprozentkennwert parameters and model output"""
    try:
        # Time constants table - matches the table in the paper
        table_zk = np.array([
        [0.300, 0.0011, 0.7950, 0.000875],
        [0.275, 0.0535, 0.7555, 0.04041],
        [0.250, 0.104, 0.7209, 0.0797],
        [0.225, 0.168, 0.6814, 0.1144],
        [0.210, 0.220, 0.6523, 0.1435],
        [0.200, 0.263, 0.6302, 0.1657],
        [0.190, 0.321, 0.6025, 0.1934],
        [0.180, 0.403, 0.5673, 0.2286],
        [0.170, 0.534, 0.5188, 0.2770],
        [0.1675, 0.590, 0.5006, 0.2953],
        [0.165, 0.661, 0.4791, 0.3167],
        [0.164, 0.699, 0.4684, 0.3274],
        [0.163, 0.744, 0.4563, 0.3395],
        [0.1625, 0.7755, 0.4482, 0.3476],
        [0.1620, 0.811, 0.4394, 0.3564],
        [0.1615, 0.860, 0.4279, 0.3680],
        [0.1610, 0.962, 0.4056, 0.3902],
        [0.160965, 1.000, 0.3979, 0.3979]
        ])

        # Gain calculation
        K = (flowrate_max - flowrate_min) / (feed_max - feed_min)
        
        # Determine initial and final values based on step direction
        initial_value = flowrate[ind_sprung]
        if step_direction == "down":
            final_value = flowrate_min
        else:
            final_value = flowrate_max
            
        # Calculate limit value (y∞)
        y_inf = final_value
        
        # Calculate the target value for T72% (based on Algorithm 2, line 2)
        # For T72%, we need to find when the response reaches 72% of the final value
        if step_direction == "down":
            # For step down, we're looking for when the value has decreased to 72% of the total change
            target_value = initial_value - 0.72 * (initial_value - final_value)
        else:
            # For step up, we're looking for when the value has increased to 72% of the total change
            target_value = initial_value + 0.72 * (final_value - initial_value)

        # Find the index where the response is closest to the 72% target
        t1_idx = np.argmin(np.abs(flowrate - target_value))
        
        # Convert to time in seconds (T1)
        timestamp_numeric = mdates.date2num(timestamps)
        T1 = (timestamp_numeric[t1_idx] - timestamp_numeric[ind_sprung]) * 24 * 3600  # seconds
        
        # Calculate T2 (based on Algorithm 2, line 6)
        T2 = 0.2847 * T1
        
        # Find the index closest to T2 seconds after the step
        t2_time = timestamp_numeric[ind_sprung] + (T2 / (24 * 3600))
        t2_idx = np.argmin(np.abs(timestamp_numeric - t2_time))
        
        # Get the y2 value at time T2
        y2 = flowrate[t2_idx]
        
        # Calculate the ratio y2/y∞ (to match the table lookup in the paper)
        if step_direction == "down":
            # For step down, we need to normalize the ratio correctly
            y2_ratio = (y2 - final_value) / (initial_value - final_value)
        else:
            # For step up
            y2_ratio = (y2 - initial_value) / (final_value - initial_value)
        
        # Find closest match in table_zk (Column 0 contains y2/y∞ values)
        idx = np.argmin(np.abs(table_zk[:, 0] - y2_ratio))
        
        # Get time constants from the table (TA/T1 and TB/T1)
        TA_ratio = table_zk[idx, 2]  # Column 2 is TA/T1
        TB_ratio = table_zk[idx, 3]  # Column 3 is TB/T1
        
        # Calculate actual time constants
        TA = TA_ratio * T1
        TB = TB_ratio * T1
        
        print(f"Time constants calculation:")
        print(f"T1 (T72%): {T1:.2f} seconds")
        print(f"T2: {T2:.2f} seconds")
        print(f"y2/y∞ ratio: {y2_ratio:.6f}")
        print(f"TA/T1: {TA_ratio:.6f}")
        print(f"TB/T1: {TB_ratio:.6f}")
        print(f"TA: {TA:.6f} seconds")
        print(f"TB: {TB:.6f} seconds")

        # Create transfer functions as in the paper: G(s) = K/((TA·s+1)·(TB·s+1))
        G1_num, G1_den = [1], [TA, 1]
        G2_num, G2_den = [1], [TB, 1]
        G_num = [K]
        G_den = [1]

        # Multiply transfer functions
        combined_num, combined_den = signal.convolve(G_num, G1_num), signal.convolve(G_den, G1_den)
        combined_num, combined_den = signal.convolve(combined_num, G2_num), signal.convolve(combined_den, G2_den)

        # Discretize transfer function (as in Algorithm 2, line 10)
        Ta = 120  # Sampling time in seconds
        Gd = signal.cont2discrete((combined_num, combined_den), Ta, method="zoh")

        # Extract transfer function coefficients for discrete form
        num = Gd[0][0]  # Get numerator coefficients
        den = Gd[1]     # Get denominator coefficients
        
        # Convert to the form Gd(z) = (e·z + f)/(z² + c·z + d) 
        e = num[0]      # Coefficient of z
        f = num[1]      # Constant term
        c = -den[1]     # Coefficient of z (note the negative sign)
        d = -den[2]     # Constant term (note the negative sign)

        # Print the coefficients for verification
        print(f"Transfer function coefficients:")
        print(f"e = {e:.6f}")
        print(f"f = {f:.6f}")
        print(f"c = {c:.6f}")
        print(f"d = {d:.6f}")

        # Simulate step response
        time_points = np.linspace(0, len(feed) - 1, len(feed))
        if step_direction == "down":
            t_input = feed - feed_max
            _, yd = signal.dlsim(Gd, t_input)  # Only two outputs
            yd = yd.flatten() + flowrate_max
        else:
            t_input = feed - feed_min
            _, yd = signal.dlsim(Gd, t_input)  # Only two outputs
            yd = yd.flatten() + flowrate_min

        # Calculate metrics
        metrics = calculate_metrics(flowrate, yd)

        # Discretize transfer function (as in Algorithm 2, line 10)
        Ta_2h = 7200  # Sampling time in seconds
        Gd_2h = signal.cont2discrete((combined_num, combined_den), Ta_2h, method="zoh")

        # Extract transfer function coefficients for discrete form
        num_2h = Gd_2h[0][0]  # Get numerator coefficients
        den_2h = Gd_2h[1]     # Get denominator coefficients
        
        # Convert to the form Gd(z) = (e·z + f)/(z² + c·z + d) 
        e_2h = num_2h[0]      # Coefficient of z
        f_2h = num_2h[1]      # Constant term
        c_2h = -den_2h[1]     # Coefficient of z (note the negative sign)
        d_2h = -den_2h[2]     # Constant term (note the negative sign)

        return {
            "model_output": yd,
            "metrics": metrics,
            "parameters": {
                "K": K,
                "TA": TA,
                "TB": TB,
                "T1": T1,
                "T2": T2,
                "y2_ratio": y2_ratio,
                "y2": y2,
                "ind_sprung": ind_sprung,
                "feed_max": feed_max,
                "feed_min": feed_min,
                "flowrate_max": flowrate_max,
                "flowrate_min": flowrate_min,
            },
            "Gd": Gd,  # Original discretized transfer function
            "transfer_function_coeffs": {
                "e": e,
                "f": f,
                "c": c,
                "d": d
            },
            "Gd_2h": Gd_2h,  # Original discretized transfer function
            "transfer_function_coeffs": {
                "e_2h": e_2h,
                "f_2h": f_2h,
                "c_2h": c_2h,
                "d_2h": d_2h
            }
        }

    except Exception as e:
        print(f"Error in calculate_zeitprozentkennwert: {e}")
        traceback.print_exc()  # Added for better error tracking
        return None

def update_model_up_plot():
    try:
        # Check if we have preprocessed data
        if global_vars['biogas_segment_up'] is None or global_vars['substrate_segment_up'] is None:
            print("No preprocessed data available. Please preprocess data first.")
            return

        # Define constants for plotting
        const = {
            'font': 'serif',
            'fontsize': 12,
            'fontsizelegend': 8,
            'fontsizeticks': 10,
            'linienbreite': 1,
            'marker_size': 4
        }

        # Define colors with enhanced organization
        colors = {
            'biogas_line': '#0000A7',      # Deep blue for biogas line
            'zeitprozent': '#EECC16',      # Yellow for time percentage line
            'substrate': '#C1272D',        # Red for substrate line
            'grid': '#E6E6E6',             # Light gray for grid
            'markers': '#AAAAAA',          # Gray for T1 and T2 markers
            'step_marker': '#000000'       # Black for step change marker
        }
            
        # Clear and set up the figure
        fig_model_up.clear()
        ax1 = fig_model_up.add_subplot(111)
        ax2 = ax1.twinx()
        
        # Get preprocessed data
        biogas_data = global_vars['biogas_segment_up']['SmoothedValueNum']
        substrate_data = global_vars['substrate_segment_up']['FeedingRate']
        timestamps = global_vars['timestamps_up']
        
        # Convert to numpy arrays for consistency
        biogas_array = np.array(biogas_data)
        substrate_array = np.array(substrate_data)
        
        # Convert timestamps to a list to ensure integer indexing works
        timestamps_list = timestamps.tolist() if hasattr(timestamps, 'tolist') else list(timestamps)

        # Process substrate data separately for plotting
        valid_substrate_mask = substrate_array > 0
        valid_substrate = substrate_array[valid_substrate_mask]
        valid_timestamps_substrate = np.array(timestamps)[valid_substrate_mask]

        # IMPROVED STEP DETECTION: Look for significant increases in substrate feeding rate
        # Calculate differences between consecutive points
        diff_substrate = np.diff(substrate_array)
        
        # For upward step, find the most positive change (largest increase)
        # and ensure it's a significant change (at least 15% of the range)
        significant_threshold = 0.15 * (np.max(substrate_array) - np.min(substrate_array))
        significant_rises = np.where(diff_substrate > significant_threshold)[0]
        
        if len(significant_rises) > 0:
            # Take the index of the largest rise
            ind_sprung = significant_rises[np.argmax(diff_substrate[significant_rises])]
        else:
            # Fallback: use the original method if no significant rises found
            ind_sprung = np.argmax(np.diff(substrate_array))
            print("WARNING: No significant substrate rise detected. Using largest change.")
            
        print(f"Step change detected at index {ind_sprung}")

        # Calculate model parameters
        feed_max = np.max(substrate_array)
        feed_min = np.min(valid_substrate) if len(valid_substrate) > 0 else 0
        flowrate_max = np.max(biogas_array)
        flowrate_min = np.min(biogas_array)

        # Calculate margins for y-axis limits
        biogas_margin = 0.05 * (flowrate_max - flowrate_min)
        substrate_margin = 0.05 * (feed_max - feed_min)
        
        print(f"Flowrate range: {flowrate_min:.2f} to {flowrate_max:.2f}")
        print(f"Feed range: {feed_min:.2f} to {feed_max:.2f}")
        
        # Set y-axis limits with margins
        ax1.set_ylim([flowrate_min - biogas_margin, flowrate_max + biogas_margin])
        ax2.set_ylim([feed_min - substrate_margin, feed_max + substrate_margin])
        
        # Calculate zeitprozentkennwert parameters with consistent data types
        zk_up = calculate_zeitprozentkennwert(
            biogas_array,
            substrate_array,
            ind_sprung,
            feed_max,
            feed_min,
            flowrate_max,  
            flowrate_min,
            timestamps,
            step_direction="up"
        )
        
        if zk_up is None:
            print("Error: Time percentage method calculation failed")
            return
            
        # Setup axes properties
        ax1.grid(True, color=colors['grid'], linestyle='-', alpha=0.3)
        ax1.xaxis.set_minor_locator(AutoMinorLocator())
        ax1.yaxis.set_minor_locator(AutoMinorLocator())
        
        # Plot biogas production rate
        line1 = ax1.plot(timestamps, biogas_array, 
                       linestyle='--',
                       color=colors['biogas_line'],
                       label='Biogas production rate',
                       linewidth=const['linienbreite'])
        
        # Plot time percentage method result - use solid line for better visibility
        line2 = ax1.plot(timestamps, zk_up['model_output'],
                       linestyle='-',  # Changed from -. to solid line
                       color=colors['zeitprozent'],
                       label='Time percentage method',
                       linewidth=const['linienbreite'])

        # Configure primary axis
        ax1.tick_params(axis='y', labelcolor=colors['biogas_line'], labelsize=const['fontsizeticks'])
        ax1.set_xlabel('Time', fontname=const['font'], fontsize=const['fontsize'])
        ax1.set_ylabel('Biogas production rate [m³/h]', 
                      fontname=const['font'], 
                      fontsize=const['fontsize'],
                      color=colors['biogas_line'])
        ax1.set_title('Time percentage method', 
                     fontname=const['font'], 
                     fontsize=const['fontsize'])

        # Configure secondary axis
        ax2.spines['right'].set_color(colors['substrate'])
        ax2.tick_params(axis='y', colors=colors['substrate'], labelsize=const['fontsizeticks'])
        ax2.yaxis.set_minor_locator(AutoMinorLocator())
        ax2.set_ylabel('Substrate feeding rate [t/h]',
                      fontname=const['font'],
                      fontsize=const['fontsize'],
                      color=colors['substrate'])
        
        # Plot substrate feeding rate
        line3 = ax2.plot(valid_timestamps_substrate, valid_substrate,
                       color=colors['substrate'],
                       linewidth=const['linienbreite'],
                       label='Substrate feeding rate')
        
        # Mark the step change point
        ax2.axvline(x=timestamps_list[ind_sprung], color=colors['step_marker'], linestyle=':', linewidth=0.8, alpha=0.7)
        ax1.text(timestamps_list[ind_sprung], flowrate_min - 0.5*biogas_margin, 
                "Step Change", fontsize=7, ha='center', va='top', rotation=90)
        
        # Format time axis
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%H:%M'))
        plt.setp(ax1.xaxis.get_majorticklabels(), 
                rotation=0,
                fontname=const['font'],
                fontsize=const['fontsizeticks'])
        
        # Add legend
        lines = line1 + line2 + line3
        labels = [l.get_label() for l in lines]
        ax1.legend(lines, labels,
                  loc='upper right',
                  fontsize=const['fontsizelegend'],
                  prop={'family': const['font']})
        
        # Add T1 and T2 markers
        if 'parameters' in zk_up:
            # Get the step time (start time)
            timestamp_numeric = mdates.date2num(timestamps)
            step_time = timestamp_numeric[ind_sprung]
            
            # Calculate time points for T1 and T2
            T1 = zk_up['parameters']['T1']  # T1 in seconds
            T2 = zk_up['parameters']['T2']  # T2 in seconds
            
            # Convert to datetime indices
            t1_time = step_time + (T1 / (24 * 3600))  # Convert seconds to days
            t2_time = step_time + (T2 / (24 * 3600))  # Convert seconds to days
            
            # Find closest indices and convert to Python int
            t1_idx = int(np.argmin(np.abs(timestamp_numeric - t1_time)))
            t2_idx = int(np.argmin(np.abs(timestamp_numeric - t2_time)))
            
            # Ensure indices are within valid range
            t1_idx = min(t1_idx, len(timestamps) - 1)
            t2_idx = min(t2_idx, len(timestamps) - 1)
            
            print(f"T1 index: {t1_idx}, max index: {len(timestamps)-1}")
            print(f"T2 index: {t2_idx}, max index: {len(timestamps)-1}")
            
            # Get y-values at T1 and T2
            y_t1 = biogas_array[t1_idx]
            y_t2 = biogas_array[t2_idx]
            
            # Get y range for plotting vertical lines
            y_min = flowrate_min - biogas_margin
            y_max = flowrate_max + biogas_margin
            
            # Plot T1 marker (72% mark)
            # Vertical line
            ax1.plot([timestamps_list[t1_idx], timestamps_list[t1_idx]], [y_min, y_t1], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Horizontal line - start from step index, not from beginning
            ax1.plot([timestamps_list[ind_sprung], timestamps_list[t1_idx]], [y_t1, y_t1], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Add T1 annotation
            ax1.text(timestamps_list[min(t1_idx+5, len(timestamps_list)-1)], y_t1, 
                   f"T1 (0.72)", fontsize=8, color=colors['markers'],
                   horizontalalignment='right', verticalalignment='bottom')
            
            # Plot T2 marker
            # Vertical line
            ax1.plot([timestamps_list[t2_idx], timestamps_list[t2_idx]], [y_min, y_t2], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Horizontal line - start from step index, not from beginning
            ax1.plot([timestamps_list[ind_sprung], timestamps_list[t2_idx]], [y_t2, y_t2], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Add T2 annotation with the y2_ratio value
            ax1.text(timestamps_list[min(t2_idx+5, len(timestamps_list)-1)], y_t2, 
                   f"T2 ({zk_up['parameters'].get('y2_ratio', 0):.3f})", 
                   fontsize=8, color=colors['markers'],
                   horizontalalignment='right', verticalalignment='bottom')
            
            # Mark the actual points with circles
            ax1.plot(timestamps_list[t1_idx], y_t1, 'o', 
                   markersize=4, markerfacecolor=colors['markers'], markeredgecolor='none')
            ax1.plot(timestamps_list[t2_idx], y_t2, 'o', 
                   markersize=4, markerfacecolor=colors['markers'], markeredgecolor='none')

        # Calculate standard metrics
        if 'model_output' in zk_up and hasattr(zk_up['model_output'], '__len__'):
            # Calculate standard metrics
            metrics = calculate_metrics(biogas_array, zk_up['model_output'])
            
            # Calculate AIC and SSE
            aic_metrics = calculate_aic_metrics(biogas_array, zk_up['model_output'])
            
            # Combine metrics
            combined_metrics = {**metrics, **aic_metrics}
            
            # Print AIC and SSE information
            print(f"======= TIME PERCENTAGE MODEL EVALUATION =======")
            print(f"AIC: {aic_metrics['AIC']:.4f}")
            print(f"AICc (corrected): {aic_metrics['AICc']:.4f}")
            print(f"SSE: {aic_metrics['SSE']:.4f}")
            print(f"MSE: {aic_metrics['MSE']:.4f}")
            print(f"RMSE: {aic_metrics['RMSE']:.4f}")
            print(f"n (number of data points): {aic_metrics['n']}")
            print(f"P (number of parameters): {aic_metrics['P']}")
            print(f"R²: {metrics.get('R2', 0):.4f}")
            print(f"==============================================")
            
            # Update metrics table with combined metrics
            update_metrics_table(tree_up, combined_metrics)

        # Store model parameters and data for control system
        global_vars["up_scenario_data"] = {
            "time": timestamps,
            "flow_measured": biogas_array,
            "model_output": zk_up["model_output"],
            "feed": substrate_array,
            # Control system parameters
            "K": zk_up['parameters']['K'],        # Process gain
            "T": zk_up['parameters']['TA'],       # Main time constant
            "T_summe": zk_up['parameters']['TA'] + zk_up['parameters']['TB'],  # Sum of time constants
            "T1": zk_up['parameters']['T1'],      # Primary time constant
            "T2": zk_up['parameters']['T2'],      # Secondary time constant
            "ind_sprung": zk_up['parameters']['ind_sprung'],  # Step index
            # Operating points
            "feed_max": feed_max,
            "feed_min": feed_min,
            "flowrate_max": flowrate_max,
            "flowrate_min": flowrate_min,
            # Additional parameters
            "initial_value": zk_up['parameters'].get('initial_value', flowrate_min),
            "final_value": zk_up['parameters'].get('final_value', flowrate_max),
            "y2_ratio": zk_up['parameters'].get('y2_ratio', 0),  # Add y2_ratio for reference
            # Model evaluation metrics
            "AIC": aic_metrics.get('AIC', float('inf')),
            "SSE": aic_metrics.get('SSE', float('inf')),
            "R2": metrics.get('R2', 0)
        }
        print("Up Scenario Data Stored with AIC and SSE metrics")
        
        # Adjust layout and draw
        fig_model_up.tight_layout()
        canvas_model_up.draw()
        
    except Exception as e:
        print(f"Error in update_model_up_plot: {e}")
        import traceback
        traceback.print_exc()


def update_model_down_plot():
    """
    Update the step down plot with time percentage method.
    Added AIC calculation to evaluate model quality.
    """
    try:
        # Check if we have preprocessed data
        if global_vars['biogas_segment_down'] is None or global_vars['substrate_segment_down'] is None:
            print("No preprocessed data available. Please preprocess data first.")
            return
            
        # Define constants
        const = {
            'font': 'serif',
            'fontsize': 12,
            'fontsizelegend': 8,
            'fontsizeticks': 10,
            'linienbreite': 1,
            'marker_size': 4
        }
        
        # Define colors
        colors = {
            'biogas_line': '#0000A7',      # Deep blue for biogas line
            'zeitprozent': '#EECC16',      # Yellow for time percentage line
            'substrate': '#C1272D',        # Red for substrate line
            'grid': '#E6E6E6',             # Light gray for grid
            'markers': '#AAAAAA',          # Gray for T1 and T2 markers
            'step_marker': '#000000'       # Black for step change marker
        }
        
        fig_model_down.clear()
        ax1 = fig_model_down.add_subplot(111)
        ax2 = ax1.twinx()
        
        # Get preprocessed data
        biogas_data = global_vars['biogas_segment_down']['SmoothedValueNum']
        substrate_data = global_vars['substrate_segment_down']['FeedingRate']
        timestamps = global_vars['timestamps_down']
        
        # Convert to numpy arrays for consistency
        biogas_array = np.array(biogas_data)
        substrate_array = np.array(substrate_data)
        
        # Convert timestamps to a list to ensure integer indexing works
        timestamps_list = timestamps.tolist() if hasattr(timestamps, 'tolist') else list(timestamps)
        
        # Process substrate data separately for plotting
        valid_substrate_mask = substrate_array > 0
        valid_substrate = substrate_array[valid_substrate_mask]
        valid_timestamps_substrate = np.array(timestamps)[valid_substrate_mask]

        # IMPROVED STEP DETECTION: Look for significant drops in substrate feeding rate
        # Calculate differences between consecutive points
        diff_substrate = np.diff(substrate_array)
        
        # For downward step, find the most negative change (largest decrease)
        # and ensure it's a significant change (at least 15% of the range)
        significant_threshold = -0.15 * (np.max(substrate_array) - np.min(substrate_array))
        significant_drops = np.where(diff_substrate < significant_threshold)[0]
        
        if len(significant_drops) > 0:
            # Take the index of the largest drop
            ind_sprung = significant_drops[np.argmin(diff_substrate[significant_drops])]
        else:
            # Fallback: use the original method if no significant drops found
            ind_sprung = np.argmax(np.abs(np.diff(substrate_array)))
            print("WARNING: No significant substrate drop detected. Using largest change.")
        
        print(f"Step change detected at index {ind_sprung}")
        
        # Calculate key values for processing
        feed_max = np.max(substrate_array)
        feed_min = np.min(valid_substrate) if len(valid_substrate) > 0 else 0
        flowrate_max = np.max(biogas_array)
        flowrate_min = np.min(biogas_array)
        
        # Calculate margin for y-axis limits
        biogas_margin = 0.05 * (flowrate_max - flowrate_min)
        substrate_margin = 0.05 * (feed_max - feed_min)

        print(f"Flowrate range: {flowrate_min:.2f} to {flowrate_max:.2f}")
        print(f"Feed range: {feed_min:.2f} to {feed_max:.2f}")
        
        # Set y-axis limits with margins
        ax1.set_ylim([flowrate_min - biogas_margin, flowrate_max + biogas_margin])
        ax2.set_ylim([feed_min - substrate_margin, feed_max + substrate_margin])
        
        # Calculate zeitprozentkennwert parameters using consistent data
        zk_down = calculate_zeitprozentkennwert(
            biogas_array,
            substrate_array,
            ind_sprung,
            feed_max,
            feed_min,
            flowrate_max,
            flowrate_min,
            timestamps,
            step_direction="down"
        )
        
        if zk_down is None:
            print("Error: Time percentage method calculation failed")
            return
            
        # Setup axes properties
        ax1.grid(True, alpha=0.3)
        ax1.xaxis.set_minor_locator(AutoMinorLocator())
        ax1.yaxis.set_minor_locator(AutoMinorLocator())
            
        # Plot biogas production rate
        line1 = ax1.plot(timestamps, biogas_array, 
                       linestyle='--',
                       color=colors['biogas_line'],
                       label='Biogas production rate',
                       linewidth=const['linienbreite'])
            
        # Plot time percentage method result
        line2 = ax1.plot(timestamps, zk_down['model_output'],
                       linestyle='-',  # Changed from -. to solid line for better visibility
                       color=colors['zeitprozent'],
                       label='Time percentage method',
                       linewidth=const['linienbreite'])
            
        # Plot substrate feeding rate (using valid data only)
        line3 = ax2.plot(valid_timestamps_substrate, valid_substrate,
                       linestyle='-',
                       color=colors['substrate'],
                       label='Substrate feeding rate',
                       linewidth=const['linienbreite'])
            
        # Configure axes
        ax1.set_xlabel('Time', fontname=const['font'], fontsize=const['fontsize'])
        ax1.set_ylabel('Biogas production rate [m³/h]',
                     fontname=const['font'],
                     fontsize=const['fontsize'],
                     color=colors['biogas_line'])
        ax2.set_ylabel('Substrate feeding rate [t/h]',
                     fontname=const['font'],
                     fontsize=const['fontsize'],
                     color=colors['substrate'])
            
        ax1.set_title('Time percentage method',
                    fontname=const['font'],
                    fontsize=const['fontsize'])
            
        # Format time axis
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%H:%M'))
        plt.setp(ax1.xaxis.get_majorticklabels(),
                rotation=0,
                fontname=const['font'],
                fontsize=const['fontsizeticks'])
            
        # Configure grid and ticks
        ax1.tick_params(axis='y', labelcolor=colors['biogas_line'], labelsize=const['fontsizeticks'])
        ax2.tick_params(axis='y', labelcolor=colors['substrate'], labelsize=const['fontsizeticks'])
            
        # Add legend
        lines = line1 + line2 + line3
        labels = [l.get_label() for l in lines]
        ax1.legend(lines, labels,
                  loc='upper right',
                  fontsize=const['fontsizelegend'],
                  prop={'family': const['font']})
        
        # Mark the step change point
        ax2.axvline(x=timestamps_list[ind_sprung], color=colors['step_marker'], linestyle=':', linewidth=0.8, alpha=0.7)
        ax1.text(timestamps_list[ind_sprung], flowrate_min - 0.5*biogas_margin, 
                "Step Change", fontsize=7, ha='center', va='top', rotation=90)
        
        # Get T1 and T2 values from zk_down
        if 'parameters' in zk_down:
            # Get the step time (start time)
            timestamp_numeric = mdates.date2num(timestamps)
            step_time = timestamp_numeric[ind_sprung]
            
            # Calculate time points for T1 and T2
            T1 = zk_down['parameters']['T1']  # T1 in seconds
            T2 = zk_down['parameters']['T2']  # T2 in seconds
            
            # Convert to datetime indices
            t1_time = step_time + (T1 / (24 * 3600))  # Convert seconds to days
            t2_time = step_time + (T2 / (24 * 3600))  # Convert seconds to days
            
            # Find closest indices and convert to Python int
            t1_idx = int(np.argmin(np.abs(timestamp_numeric - t1_time)))
            t2_idx = int(np.argmin(np.abs(timestamp_numeric - t2_time)))
            
            # Ensure indices are within valid range
            t1_idx = min(t1_idx, len(timestamps) - 1)
            t2_idx = min(t2_idx, len(timestamps) - 1)
            
            print(f"T1 index: {t1_idx}, max index: {len(timestamps)-1}")
            print(f"T2 index: {t2_idx}, max index: {len(timestamps)-1}")
            
            # Get y-values at T1 and T2
            y_t1 = biogas_array[t1_idx]
            y_t2 = biogas_array[t2_idx]
            
            # Get y range for plotting vertical lines
            y_min = flowrate_min - biogas_margin
            y_max = flowrate_max + biogas_margin
            
            # Plot T1 marker (72% mark)
            # Vertical line
            ax1.plot([timestamps_list[t1_idx], timestamps_list[t1_idx]], [y_min, y_t1], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Horizontal line - start from step index, not from beginning
            ax1.plot([timestamps_list[ind_sprung], timestamps_list[t1_idx]], [y_t1, y_t1], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Add T1 annotation
            ax1.text(timestamps_list[min(t1_idx+5, len(timestamps_list)-1)], y_t1, 
                   f"T1 (0.72)", fontsize=8, color=colors['markers'],
                   horizontalalignment='right', verticalalignment='bottom')
            
            # Plot T2 marker
            # Vertical line
            ax1.plot([timestamps_list[t2_idx], timestamps_list[t2_idx]], [y_min, y_t2], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Horizontal line - start from step index, not from beginning
            ax1.plot([timestamps_list[ind_sprung], timestamps_list[t2_idx]], [y_t2, y_t2], 
                   linestyle=':', color=colors['markers'], linewidth=1.0)
            # Add T2 annotation with the y2_ratio value from the calculation
            ax1.text(timestamps_list[min(t2_idx+5, len(timestamps_list)-1)], y_t2, 
                   f"T2 ({zk_down['parameters'].get('y2_ratio', 0):.3f})", 
                   fontsize=8, color=colors['markers'],
                   horizontalalignment='right', verticalalignment='bottom')
            
            # Mark the actual points with circles
            ax1.plot(timestamps_list[t1_idx], y_t1, 'o', 
                   markersize=4, markerfacecolor=colors['markers'], markeredgecolor='none')
            ax1.plot(timestamps_list[t2_idx], y_t2, 'o', 
                   markersize=4, markerfacecolor=colors['markers'], markeredgecolor='none')
            
        # Calculate standard metrics and AIC metrics
        if 'model_output' in zk_down and hasattr(zk_down['model_output'], '__len__'):
            # Calculate standard metrics
            metrics = calculate_metrics(biogas_array, zk_down['model_output'])
            
            # Calculate AIC and SSE
            aic_metrics = calculate_aic_metrics(biogas_array, zk_down['model_output'])
            
            # Combine metrics
            combined_metrics = {**metrics, **aic_metrics}
            
            # # Print AIC and SSE information
            # print(f"======= TIME PERCENTAGE MODEL EVALUATION (DOWN) =======")
            # print(f"AIC: {aic_metrics['AIC']:.4f}")
            # print(f"AICc (corrected): {aic_metrics['AICc']:.4f}")
            # print(f"SSE: {aic_metrics['SSE']:.4f}")
            # print(f"MSE: {aic_metrics['MSE']:.4f}")
            # print(f"RMSE: {aic_metrics['RMSE']:.4f}")
            # print(f"n (number of data points): {aic_metrics['n']}")
            # print(f"P (number of parameters): {aic_metrics['P']}")
            # print(f"R²: {metrics.get('R2', 0):.4f}")
            # print(f"=================================================")
            
            # Update metrics table with combined metrics
            update_metrics_table(tree_down, combined_metrics)

        # Store data for later use if needed
        global_vars["down_scenario_data"] = {
            "time": timestamps,
            "flow_measured": biogas_array,
            "model_output": zk_down["model_output"],
            "feed": substrate_array,
            # Additional model parameters
            "K": zk_down['parameters']['K'],        # Process gain
            "TA": zk_down['parameters']['TA'],      # First time constant 
            "TB": zk_down['parameters']['TB'],      # Second time constant
            "T1": zk_down['parameters']['T1'],      # Primary time constant
            "T2": zk_down['parameters']['T2'],      # Secondary time constant
            "ind_sprung": ind_sprung,               # Step index
            # Operating points
            "feed_max": feed_max,
            "feed_min": feed_min,
            "flowrate_max": flowrate_max,
            "flowrate_min": flowrate_min,
            # Additional parameters
            "initial_value": zk_down['parameters'].get('initial_value', flowrate_max),
            "final_value": zk_down['parameters'].get('final_value', flowrate_min),
            "y2_ratio": zk_down['parameters'].get('y2_ratio', 0),
            # Model evaluation metrics
            "AIC": aic_metrics.get('AIC', float('inf')),
            "SSE": aic_metrics.get('SSE', float('inf')),
            "R2": metrics.get('R2', 0)
        }
        print("Down Scenario Data Stored with AIC and SSE metrics")
            
        # Adjust layout and draw
        fig_model_down.tight_layout()
        canvas_model_down.draw()
            
    except Exception as e:
        print(f"Error in update_model_down_plot: {e}")
        import traceback
        traceback.print_exc()


def plot_characteristic_point(ax, timestamps, data, idx, value, y_min, color):
    """Helper function for plotting characteristic points with enhanced markers"""
    try:
        # Horizontal markers
        x_horiz = [timestamps[0], 
                  timestamps[round((1 + idx) / 2)],
                  timestamps[idx]]
        y_horiz = [data[idx]] * 3
        ax.plot(x_horiz, y_horiz,
               color=color,
               linewidth=1,
               linestyle=':',
               marker='>',
               markersize=4)
        
        # Vertical markers
        x_vert = [timestamps[idx]] * 3
        y_vert = [y_min + 3,
                 (y_min + data[idx]) / 2,
                 data[idx]]
        ax.plot(x_vert, y_vert,
               color=color,
               linewidth=1,
               linestyle=':',
               marker='v',
               markersize=4)
        
        # Point marker
        ax.plot(x_vert[-1], y_vert[-1], 'o',
               markersize=4,
               markeredgecolor=color,
               markerfacecolor=color)
        
        # Text annotation
        ax.text(timestamps[30],
               data[idx] + 2,
               value,
               fontsize=8,
               color=color)
    except Exception as e:
        print(f"Error in plot_characteristic_point: {e}")


def update_model_estimation_plots():
    """Update both step up and step down plots in model estimation tab based on selected model"""
    try:
        # Get selected model
        model = selected_model.get()
        
        # Clear existing plots
        fig_model_down.clear()
        fig_model_up.clear()
        
        if model == "model1":  # Time percentage
            # Use existing time percentage calculation
            update_model_down_plot()
            update_model_up_plot()
    except Exception as e:
        print(f"Error updating model plots: {e}")

def clear_plots():
    """Clear both plots and reset them to blank state"""
    try:
        # Clear down plot
        fig_model_down.clear()
        ax1_down = fig_model_down.add_subplot(111)
        ax1_down.set_xlabel('Time')
        ax1_down.set_ylabel('Biogas production rate [m³/h]')
        ax1_down.set_title('Model Output')
        fig_model_down.tight_layout()
        canvas_model_down.draw()
        
        # Clear up plot
        fig_model_up.clear()
        ax1_up = fig_model_up.add_subplot(111)
        ax1_up.set_xlabel('Time')
        ax1_up.set_ylabel('Biogas production rate [m³/h]')
        ax1_up.set_title('Model Output')
        fig_model_up.tight_layout()
        canvas_model_up.draw()
        
        # Clear metrics tables
        for tree in [tree_down, tree_up]:
            for item in tree.get_children():
                tree.delete(item)
                
    except Exception as e:
        print(f"Error clearing plots: {e}")

# Modify the radio button command to update plots when selection changes
def on_model_select():
    """Handle radio button selection change"""
    try:
        update_model_estimation_plots()
    except Exception as e:
        print(f"Error in model selection: {e}")

# Update radio button creation to include command
for i, (text, value) in enumerate(models):
    rb = tk.Radiobutton(
        radio_frame,
        text=text,
        variable=selected_model,
        value=value,
        font=("Arial", 10),
        command=on_model_select  # Add command to handle selection
    )
    rb.grid(row=0, column=i, padx=10)


# Call clear_plots initially to start with blank plots
clear_plots()

def select_model():
    """Handle model selection button click:
       1. Store the selected model's output data.
       2. Switch to the 'Control System' tab.
    """
    try:
        update_model_estimation_plots()  # Update plots with selected model

        # Get which model is selected
        model = selected_model.get()
        print(f"Selected model: {model}")

        # Ensure selected_model_data is a dictionary
        global selected_model_data  
        if 'selected_model_data' not in globals():
            selected_model_data = {}

        # Store the upward response data - using up_scenario_data that's already populated
        if "up_scenario_data" in global_vars:
            # Validate the required data is present
            required_params = ['K', 'T', 'T_summe']
            if all(param in global_vars["up_scenario_data"] for param in required_params):
                # NEW CODE: Recalculate and store transfer function coefficients
                biogas_array = np.array(global_vars['biogas_segment_up']['SmoothedValueNum'])
                substrate_array = np.array(global_vars['substrate_segment_up']['FeedingRate']) 
                timestamps = global_vars['timestamps_up']
                
                feed_max = global_vars["up_scenario_data"]["feed_max"]
                feed_min = global_vars["up_scenario_data"]["feed_min"]
                flowrate_max = global_vars["up_scenario_data"]["flowrate_max"]
                flowrate_min = global_vars["up_scenario_data"]["flowrate_min"]
                ind_sprung = global_vars["up_scenario_data"]["ind_sprung"]
                
                # Recalculate zeitprozentkennwert
                zk_up = calculate_zeitprozentkennwert(
                    biogas_array,
                    substrate_array, 
                    ind_sprung,
                    feed_max,
                    feed_min,
                    flowrate_max,
                    flowrate_min,
                    timestamps,
                    step_direction="up"
                )
                
                if zk_up and 'transfer_function_coeffs' in zk_up:
                    # Store in control_system_params for later access
                    global_vars['control_system_params'] = {
                        'transfer_function_coeffs': zk_up['transfer_function_coeffs']
                    }
                    print("Transfer function coefficients stored successfully")
                    print("Stored coefficients:", global_vars['control_system_params'])
                else:
                    print("Error: Could not calculate transfer function coefficients")
                
                # Original code continues
                selected_model_data['upward'] = global_vars["up_scenario_data"]
                print(f"Stored Upward Model Data: {selected_model_data['upward']}")
                
                # Switch to Control System tab
                notebook.select(control_tab)
            else:
                messagebox.showerror("Error", "Missing required model parameters. Please ensure model estimation is complete.")
                print("Error: Missing required parameters in up_scenario_data")
                return
        else:
            messagebox.showerror("Error", "No upward scenario data available. Please process data first.")
            print("Warning: No upward scenario data available. Please process data first.")
            return

    except Exception as e:
        messagebox.showerror("Error", f"Error in model selection: {str(e)}")
        print(f"Error in select_model: {e}")
        traceback.print_exc()


# Add hover effects for buttons
def on_enter_model(event):
    event.widget['background'] = '#d0eaff'  # Light blue color on hover 

def on_leave_model(event):
    event.widget['background'] = 'SystemButtonFace'  # Default button color

# Bind hover effects to buttons
model_estimation_button.bind("<Enter>", on_enter_model)
model_estimation_button.bind("<Leave>", on_leave_model)
select_model_button.bind("<Enter>", on_enter_model)
select_model_button.bind("<Leave>", on_leave_model)

###################################################################################################
# CONTROL SYSTEMS
# Configure the Control Systems tab
control_tab.grid_rowconfigure(0, weight=1)
control_tab.grid_columnconfigure(0, weight=1)

# Main frame for the Control Systems section
main_frame = ttk.Frame(control_tab, padding=20)
main_frame.grid(row=0, column=0, sticky='nsew')
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_rowconfigure(2, weight=1)

# Top frame for inputs
top_frame = ttk.Frame(main_frame, padding=10)
top_frame.grid(row=0, column=0, sticky='ew', pady=10)
top_frame.grid_columnconfigure(0, weight=3)
top_frame.grid_columnconfigure(1, weight=2)

# Feedback Controller Section
feedback_frame = ttk.LabelFrame(top_frame, text="Feedback Controller", padding=10)
feedback_frame.grid(row=0, column=0, sticky='nsew', padx=10)
feedback_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

def create_input(frame, label, unit, row, column):
    """Helper to create a labeled input field"""
    input_frame = ttk.Frame(frame)
    input_frame.grid(row=row, column=column, sticky='w', padx=5, pady=5)
    ttk.Label(input_frame, text=label, style='InputLabel.TLabel').pack(side='left', padx=(0, 10))
    entry = ttk.Entry(input_frame, width=15)
    entry.pack(side='left', padx=(0, 5))
    ttk.Label(input_frame, text=unit, style='InputLabel.TLabel').pack(side='left')
    return entry

# Create input fields with default values
gas_flow_initial_field = create_input(feedback_frame, "Gas production flow rate, initial state", "[m³/h]", 0, 0)
gas_flow_initial_field.insert(0, "83")

gas_flow_setpoint_field = create_input(feedback_frame, "Gas production flow rate, setpoint", "[m³/h]", 0, 2)
gas_flow_setpoint_field.insert(0, "95")

substrate_feed_initial_field = create_input(feedback_frame, "Substrate feeding rate, initial state", "[t/2h]", 1, 0)
substrate_feed_initial_field.insert(0, "0.42")

start_date_field = DateEntry(feedback_frame, width=12, background='darkblue', foreground='white', borderwidth=2)
start_date_label = ttk.Label(feedback_frame, text="Start date")
start_date_label.grid(row=2, column=0, padx=5, pady=5, sticky='e')
start_date_field.grid(row=2, column=1, padx=5, pady=5, sticky='w')

duration_field = create_input(feedback_frame, "Simulation duration", "[days]", 2, 2)
duration_field.insert(0, "2")

# Saturations Section
saturations_frame = ttk.LabelFrame(top_frame, text="Saturations", padding=10)
saturations_frame.grid(row=0, column=1, sticky='nsew', padx=10)
saturations_frame.grid_columnconfigure(0, weight=1)

feed_max_field = create_input(saturations_frame, "Substrate feeding rate (max)", "[t/2h]", 0, 0)
feed_max_field.insert(0, "0.8")

feed_min_field = create_input(saturations_frame, "Substrate feeding rate (min)", "[t/2h]", 1, 0)
feed_min_field.insert(0, "0.4025")

# Plot Section
plot_frame = ttk.Frame(main_frame, padding=10)
plot_frame.grid(row=2, column=0, sticky='nsew')
plot_frame.grid_rowconfigure(0, weight=1)
plot_frame.grid_columnconfigure(0, weight=1)

# Create Matplotlib figure and axes
fig = Figure(figsize=(8, 4))
ax = fig.add_subplot(111)
ax.set_title("Discrete PI-Controller (+ Root Locus) with Anti-Windup")

# Create canvas and embed in plot_frame
canvas = FigureCanvasTkAgg(fig, master=plot_frame)
canvas.get_tk_widget().grid(row=0, column=0, sticky='nsew')

# Create button frame
button_frame = ttk.Frame(feedback_frame)
button_frame.grid(row=3, column=0, columnspan=4, pady=10)

def get_ui_values():
    """
    Retrieve all current values from the UI fields and store them in a dictionary.
    This function can be called whenever we need the latest values from the UI.
    """
    ui_values = {
        "gas_flow_initial": float(gas_flow_initial_field.get()),
        "gas_flow_setpoint": float(gas_flow_setpoint_field.get()),
        "substrate_feed_initial": float(substrate_feed_initial_field.get()),
        "simulation_duration": float(duration_field.get()),
        "feed_max": float(feed_max_field.get()),
        "feed_min": float(feed_min_field.get()),
        "start_date": start_date_field.get_date()
    }
    return ui_values

def set_system_parameters():
    """
    Set the system parameters including sampling time
    """
    # Set sampling time to 7200 seconds (2 hours)
    Ts_2h = 7200
    return Ts_2h

def extract_transfer_function_coefficients():
    """
    Extract numerator and denominator coefficients from the transfer function
    using the time percentage method results from the preprocessing tab.
    """
    try:
        # Check if time percentage method results exist
        if global_vars.get('biogas_segment_up') is None or global_vars.get('control_system_params') is None:
            print("Error: Time percentage method results not available. Process data in preprocessing tab first.")
            return None
        
        # Get control system parameters from the global variables
        control_params = global_vars.get('control_system_params')
        
        # Extract transfer function coefficients
        if 'transfer_function_coeffs' in control_params:
            coeffs = control_params['transfer_function_coeffs']
            e_2h, f_2h, c_2h, d_2h = coeffs["e_2h"], coeffs["f_2h"], coeffs["c_2h"], coeffs["d_2h"]
            
            # For a discrete transfer function of the form Gd(z) = (e·z + f)/(z² + c·z + d)
            pt2_Gd_num = np.array([e_2h, f_2h])    # Numerator coefficients
            pt2_Gd_den = np.array([1, -c_2h, -d_2h])  # Denominator coefficients with leading 1
            
            print("Successfully extracted transfer function coefficients:")
            print(f"Numerator coefficients: {pt2_Gd_num}")
            print(f"Denominator coefficients: {pt2_Gd_den}")
            
            return {
                'numerator': pt2_Gd_num,
                'denominator': pt2_Gd_den
            }
        else:
            print("Error: Transfer function coefficients not found in control system parameters.")
            return None
            
    except Exception as e:
        print(f"Error extracting transfer function coefficients: {e}")
        traceback.print_exc()
        return None
    
def store_transfer_function_coefficients():
    """
    Extract and store the transfer function coefficients from the time percentage method 
    for later use in the simulation.
    """
    try:
        # Get coefficients using the extraction function
        coeffs = extract_transfer_function_coefficients()
        
        if coeffs is None:
            return False
        
        # Store coefficients in global variables for later access
        global_vars['pt2_Gd_num'] = coeffs['numerator']
        global_vars['pt2_Gd_den'] = coeffs['denominator']
        
        print("Transfer function coefficients stored successfully.")
        return True
            
    except Exception as e:
        print(f"Error storing transfer function coefficients: {e}")
        traceback.print_exc()
        return False
    
def resample_transfer_function():
    """Simple direct approach to create a transfer function with 2-hour sampling"""
    try:
        # Get the original transfer function coefficients
        if 'pt2_Gd_num' not in global_vars or 'pt2_Gd_den' not in global_vars:
            print("Error: Transfer function coefficients not found.")
            return None
            
        # Get the original coefficients
        num_original = global_vars['pt2_Gd_num']
        den_original = global_vars['pt2_Gd_den']
        
        # Create a transfer function with 2-hour sampling
        # This is a simplified approach - we'll use the same coefficients
        # For a proper resampling, a more complex approach would be needed
        
        # Store the results
        global_vars['G_2h_num'] = num_original
        global_vars['G_2h_den'] = den_original
        global_vars['G_2h'] = ctrl.TransferFunction(num_original, den_original, dt=7200)
        
        print("Created transfer function for 2-hour sampling.")
        print(f"Numerator: {global_vars['G_2h_num']}")
        print(f"Denominator: {global_vars['G_2h_den']}")
        
        # Set the original sampling time for reference
        global_vars['original_Ts'] = 120
        
        return global_vars['G_2h']
    
    except Exception as e:
        print(f"Error in resample_transfer_function: {e}")
        traceback.print_exc()
        return None

def convert_to_state_space():
    """
    Convert the discrete transfer function to a state-space representation
    using the control.tf2ss function.
    """
    try:
        # Check if the resampled transfer function exists
        if 'G_2h_num' not in global_vars or 'G_2h_den' not in global_vars:
            print("Error: Resampled transfer function not found. Run resample_transfer_function first.")
            return None
        
        # Get the resampled transfer function coefficients
        G_2h_num = global_vars['G_2h_num']
        G_2h_den = global_vars['G_2h_den']
        
        # Create a transfer function object with 2-hour sampling time
        tf_sys = ctrl.TransferFunction(G_2h_num, G_2h_den, dt=7200)
        
        # Convert to state-space representation
        ss_sys = ctrl.tf2ss(tf_sys)
        
        # Store the state-space matrices
        global_vars['A'] = ss_sys.A
        global_vars['B'] = ss_sys.B
        global_vars['C'] = ss_sys.C
        global_vars['D'] = ss_sys.D
        global_vars['ss_sys'] = ss_sys
        
        print("Transfer function converted to state-space representation successfully.")
        print(f"A matrix:\n{ss_sys.A}")
        print(f"B matrix:\n{ss_sys.B}")
        print(f"C matrix:\n{ss_sys.C}")
        print(f"D matrix:\n{ss_sys.D}")
        
        return ss_sys
        
    except Exception as e:
        print(f"Error converting to state-space: {e}")
        traceback.print_exc()
        return None
    
from control import acker

def design_deadbeat_controller():
    """
    Design a deadbeat controller using Ackermann's formula instead of direct pole placement.
    """
    try:
        # Check if state-space matrices exist
        if 'A' not in global_vars or 'B' not in global_vars:
            print("Error: State-space matrices not found.")
            return None
        
        # Get the state-space matrices
        A = global_vars['A']
        B = global_vars['B']
        
        # Check system dimensions
        system_order = A.shape[0]
        
        # For a 2nd order system (as in the paper)
        if system_order == 2:
            try:
                # Try exact zeros for true deadbeat response
                desired_poles = np.array([0.0, 0.0])  # Exact zeros
                print(f"Attempting to place poles at: {desired_poles}")
                
                # Use control library's acker function (Ackermann's formula)
                K = ctrl.acker(A, B, desired_poles)
                
                # Check if K was successfully calculated
                if K is not None and K.size > 0:
                    global_vars['K'] = K
                    print(f"Deadbeat controller designed successfully using Ackermann's formula: {K}")
                    return K
                else:
                    raise ValueError("Failed to calculate K with exact zeros using Ackermann's formula")
                    
            except Exception as e1:
                print(f"Exact pole placement with Ackermann's formula failed: {e1}")
                
                try:
                    # Try with very small values near zero
                    desired_poles = np.array([0.009, 0.008])  # Very close to zero
                    print(f"Trying near-zero poles at: {desired_poles}")
                    K = ctrl.acker(A, B, desired_poles)
                    
                    global_vars['K'] = K
                    print(f"Near-zero controller designed using Ackermann's formula: {K}")
                    return K
                    
                except Exception as e2:
                    print(f"Near-zero pole placement with Ackermann's formula failed: {e2}")
                    
                    # Last resort: Calculate K manually for this specific system
                    K = np.zeros((1, 2))
                    # For a deadbeat response, we want eigenvalues of (A-BK) to be 0
                    # For the specific form of A and B in this system:
                    K[0, 0] = A[0, 0]  # First coefficient of A
                    K[0, 1] = A[0, 1]  # Second coefficient of A
                    
                    global_vars['K'] = K
                    print(f"Using manually calculated controller gain: {K}")
                    return K
        else:
            # For other system orders, try Ackermann's formula first
            try:
                desired_poles = np.zeros(system_order)  # All zeros for true deadbeat
                K = ctrl.acker(A, B, desired_poles)
                global_vars['K'] = K
                print(f"Using Ackermann's formula for system of order {system_order}: {K}")
                return K
            except Exception as e:
                print(f"Ackermann's formula failed for system order {system_order}: {e}")
                
                # Fallback to a simple approach
                K = np.zeros((1, system_order))
                for i in range(system_order):
                    K[0, i] = A[0, i] if i < A.shape[1] else 0.0
                
                global_vars['K'] = K
                print(f"Using system-based controller gain: {K}")
                return K
            
    except Exception as e:
        print(f"Error designing controller: {e}")
        traceback.print_exc()
        
        # Fallback controller that at least matches the system structure
        system_order = global_vars['A'].shape[0]
        K = np.zeros((1, system_order))
        K[0, 0] = 0.95
        K[0, 1] = 0.5 if system_order > 1 else 0.0
        
        global_vars['K'] = K
        print(f"Using fallback controller gain: {K}")
        return K


def calculate_prefilter_gain():
    """
    Calculate prefilter gain using the exact formula from MATLAB:
    F = [C·(I - A + B·K)^(-1)·B]^(-1)
    """
    # Get matrices
    A = global_vars['A']
    B = global_vars['B']
    C = global_vars['C']
    K = global_vars['K']
    
    # Create identity matrix of appropriate size
    I = np.eye(A.shape[0])
    
    # Calculate (I - A + B·K)
    BK = B @ K
    inner_term = I - A + BK
    
    # Calculate (I - A + B·K)^(-1)
    inner_inverse = np.linalg.inv(inner_term)
    
    # Calculate C·(I - A + B·K)^(-1)·B
    C_inner_B = C @ inner_inverse @ B
    
    # Calculate [C·(I - A + B·K)^(-1)·B]^(-1)
    F = 1.0 / C_inner_B.item()
    
    global_vars['f_ss_PT2'] = F
    return F


###########################

def plot_with_antiwindup(simulation_results, ui_values):
    """
    Plot simulation results with anti-windup styling
    Based on the MATLAB function plot_regler_simulink_anti_windup_ui
    """
    # Clear the figure
    fig.clear()
    
    # Create primary axis
    ax1 = fig.add_subplot(111)
    
    # Create secondary axis for substrate feeding
    ax2 = ax1.twinx()
    
    # Get simulation results
    time = simulation_results['time']  # Time in hours
    biogas = simulation_results['biogas']  # Biogas production
    substrate_feeding = simulation_results['substrate_feeding']  # Substrate feeding
    
    # Convert simulation start time to datetime
    start_date = ui_values["start_date"]
    start_datetime = pd.to_datetime(start_date)
    
    # Define color scheme similar to MATLAB version
    cmap = ['#0072BD', '#D95319', '#EDB120', '#7E2F8E', '#77AC30', '#4DBEEE', '#A2142F']
    
    # Font and styling parameters
    const = {
        'font': 'serif',
        'fontsize': 16,
        'fontsizelegend': 10,
        'fontsizeticks': 10,
        'linienbreite': 1.5,  # Line width
    }
    
    # Calculate time values with proper datetime format
    if isinstance(start_datetime, pd.Timestamp):
        datetime_values = [start_datetime + pd.Timedelta(hours=t) for t in time]
    else:
        datetime_values = time  # Fallback to numeric time values
    
    # Calculate y-axis limits with 5% padding
    y_min = min(biogas) - 0.05 * max(biogas)
    y_max = max(biogas) + 0.05 * max(biogas)
    
    # Set up primary y-axis (biogas)
    ax1.set_ylim([y_min, y_max])
    ax1.grid(True)
    ax1.xaxis.set_minor_locator(AutoMinorLocator())
    ax1.yaxis.set_minor_locator(AutoMinorLocator())
    
    # Format axis labels
    ax1.set_ylabel('Gas production flow rate [m³/h]', 
                  fontname=const['font'], 
                  fontsize=const['fontsize'],
                  color=cmap[0])
    
    if isinstance(datetime_values[0], (pd.Timestamp, datetime)):
        # Use date formatting if we have datetime objects
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d\n%H:%M'))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=0, ha='center')
        ax1.set_xlabel('Time', fontname=const['font'], fontsize=const['fontsize'])
    else:
        # Use hours formatting otherwise
        ax1.set_xlabel('Time [h]', fontname=const['font'], fontsize=const['fontsize'])
    
    # Plot biogas production on primary y-axis
    line1 = ax1.plot(datetime_values, biogas, 
                   color=cmap[0], 
                   linewidth=const['linienbreite'],
                   label='Simulated Gas production flow rate')
    
    # Plot setpoint on primary y-axis
    setpoint = ui_values['gas_flow_setpoint']
    line2 = ax1.axhline(y=setpoint, 
                      color=cmap[2], 
                      linewidth=const['linienbreite'],
                      linestyle=':', 
                      label='Setpoint',
                      alpha=0.8)
    
    # Set up secondary y-axis (substrate feeding)
    y_min_right = min(substrate_feeding) - 0.05 * max(substrate_feeding)
    y_max_right = max(substrate_feeding) + 0.05 * max(substrate_feeding)
    ax2.set_ylim([y_min_right, y_max_right])
    ax2.yaxis.set_minor_locator(AutoMinorLocator())
    
    # Set sampling time (Ta) - assume 2 hours
    Ta = 2
    ax2.set_ylabel(f'Substrate feeding rate [t/{Ta}h]',
                  fontname=const['font'],
                  fontsize=const['fontsize'],
                  color=cmap[1])
    
    # Plot substrate feeding on secondary y-axis
    line3 = ax2.step(datetime_values, substrate_feeding, 
                   where='post',
                   color=cmap[1], 
                   linewidth=const['linienbreite'],
                   label='Substrate feeding rate')
    
    # Set title to match MATLAB version
    ax1.set_title("Discrete PI-Controller (+ Root Locus) with Anti-Windup",
                 fontname=const['font'],
                 fontsize=const['fontsize'])
    
    # Create combined legend
    lines = line1 + [line2] + line3
    labels = [l.get_label() for l in lines]
    ax1.legend(lines, labels, 
              loc='lower right',
              fontsize=const['fontsizelegend'],
              prop={'family': const['font']})
    
    # Configure tick parameters
    ax1.tick_params(axis='y', labelcolor=cmap[0], labelsize=const['fontsizeticks'])
    ax2.tick_params(axis='y', labelcolor=cmap[1], labelsize=const['fontsizeticks'])
    
    # Adjust layout
    fig.tight_layout()

def run_simulation():
    """
    Run a state space controller simulation according to Fig. 7.
    Implements feedback control loop using the state space controller.
    """
    try:
        # Check each parameter individually and print status
        print("\n[DEBUG] Checking required parameters:")
        
        has_A = 'A' in global_vars
        has_B = 'B' in global_vars  
        has_C = 'C' in global_vars
        has_K = 'K' in global_vars
        has_f = 'f_ss_PT2' in global_vars
        
        print(f"A matrix exists: {has_A}")
        print(f"B matrix exists: {has_B}")
        print(f"C matrix exists: {has_C}")
        print(f"K matrix exists: {has_K}")
        print(f"f_ss_PT2 (prefilter gain) exists: {has_f}")
        
        params_exist = has_A and has_B and has_C and has_K and has_f
        
        if not params_exist:
            print("\n[DEBUG] Some parameters missing. Running prerequisite steps...")

            # Show the current state of global_vars
            print("\nCurrent global_vars keys:")
            print(list(global_vars.keys()))

            # Run prerequisites
            success_tf = store_transfer_function_coefficients()
            print(f"store_transfer_function_coefficients returned: {success_tf}")
            
            G_2h = resample_transfer_function()
            # print(f"resample_transfer_function returned G_2h: {G_2h is not None}")
            
            ss_sys = convert_to_state_space()
            print(f"convert_to_state_space returned ss_sys: {ss_sys is not None}")
            
            # Make sure controller design completed
            try:
                K = design_deadbeat_controller()
                print(f"design_deadbeat_controller returned K: {K is not None}")
            except Exception as e:
                print(f"Error designing controller: {e}")
                # Set a default K value if needed
                if 'K' not in global_vars:
                    global_vars['K'] = np.array([[0.5, 0.5]])
                    print("Using default controller gain")
            
            # Make sure prefilter gain is calculated
            try:
                F = calculate_prefilter_gain()
                print(f"calculate_prefilter_gain returned F: {F}")
            except Exception as e:
                print(f"Error calculating prefilter: {e}")
                # Set a default value if needed
                if 'f_ss_PT2' not in global_vars:
                    global_vars['f_ss_PT2'] = 1.0
                    print("Using default prefilter gain of 1.0")

            # Check again after running prerequisites
            print("\n[DEBUG] After prerequisites, checking parameters again:")
            print(f"A matrix exists: {'A' in global_vars}")
            print(f"B matrix exists: {'B' in global_vars}")
            print(f"C matrix exists: {'C' in global_vars}")
            print(f"K matrix exists: {'K' in global_vars}")
            print(f"f_ss_PT2 exists: {'f_ss_PT2' in global_vars}")
            
            # Show updated global_vars
            print("\nUpdated global_vars keys:")
            print(list(global_vars.keys()))

        # Extract parameters - using proper variable names
        A = global_vars.get('A')  # G_ss_PT2_A'vec
        B = global_vars.get('B')  # G_ss_PT2_B'u
        C = global_vars.get('C')  # G_ss_PT2_C'vec
        K = global_vars.get('K')  # K_ss_PT2'vec
        F = global_vars.get('f_ss_PT2')  # f_ss_PT2'u

        # Print debug info for all parameters
        print("\n[DEBUG] System Matrices and Controller Values:")
        print(f"A Matrix: {'None' if A is None else A.shape}")
        if A is not None: print(A)
        
        print(f"B Matrix: {'None' if B is None else B.shape}")
        if B is not None: print(B)
        
        print(f"C Matrix: {'None' if C is None else C.shape}")
        if C is not None: print(C)
        
        print(f"K (Gain Matrix): {'None' if K is None else K.shape}")
        if K is not None: print(K)
        
        print(f"F (Prefilter Gain): {F}")

        # Check for None values individually
        if A is None: 
            print("A matrix is None, using default")
            global_vars['A'] = np.array([[0.9]])
            A = global_vars['A']
            
        if B is None:
            print("B matrix is None, using default")
            global_vars['B'] = np.array([[1.0]])
            B = global_vars['B']
            
        if C is None:
            print("C matrix is None, using default")
            global_vars['C'] = np.array([[1.0]])
            C = global_vars['C']
            
        if K is None:
            print("K matrix is None, using default")
            global_vars['K'] = np.array([[0.5]])
            K = global_vars['K']
            
        if F is None:
            print("F (prefilter gain) is None, using default")
            global_vars['f_ss_PT2'] = 1.0
            F = global_vars['f_ss_PT2']

        # Print the final values we'll use for simulation
        print("\n[DEBUG] Final parameter values for simulation:")
        print(f"A: {A}")
        print(f"B: {B}")
        print(f"C: {C}")
        print(f"K: {K}")
        print(f"F: {F}")

        # Get UI values
        ui_values = get_ui_values()

        # Simulation parameters
        days = ui_values["simulation_duration"]
        Ts = set_system_parameters()  # 7200 seconds (2-hour sampling time)
        samples = int((days * 24 * 3600) / Ts)  # Number of simulation steps

        # Initialize arrays for simulation
        time = np.arange(0, samples) * Ts / 3600  # Time in hours
        biogas = np.zeros(samples)  # Biogas production (output)
        substrate_feeding = np.zeros(samples)  # Substrate feeding rate (control input)

        # Initial conditions
        biogas[0] = ui_values["gas_flow_initial"]
        biogas[1] = ui_values["gas_flow_initial"]
        #feed initial (input source:UI field)
        substrate_feeding[0] = ui_values["substrate_feed_initial"]

        # Initialize state vector
        state_dim = A.shape[0]
        x = np.zeros((samples, state_dim))  # Store all states (delta_x_koh)

        # Setpoint and control limits
        delta_setpoint = ui_values["gas_flow_setpoint"] - ui_values["gas_flow_initial"]
        feed_min = ui_values["feed_min"]
        feed_max = ui_values["feed_max"]

        # Debugging: Print simulation parameters
        print(f"\n[DEBUG] Simulation Setup:")
        print(f"Simulation Duration (days): {days}")
        print(f"Sampling Time (s): {Ts}")
        print(f"Samples: {samples}")
        print(f"Setpoint: {delta_setpoint} m³/h")
        print(f"Feed Limits: Min = {feed_min}, Max = {feed_max}")

        # Anti-windup variables
        u_unbounded = np.zeros(samples)  # Unbounded control signals
        u_bounded = np.zeros(samples)    # Bounded control signals
        u_bounded[0] = substrate_feeding[0] - ui_values["substrate_feed_initial"]

        # Simulation loop 
        for k in range(1, samples):
            # Get state from previous step (represents the Delay 1/2 block)
            state_vector = x[k-1, :].reshape(-1, 1)
            
            # System Gain K (state feedback)
            state_feedback = float((K @ state_vector)[0, 0])
            
            # Prefilter F (applies to delta_setpoint)
            prefilter_output = F * delta_setpoint
            
            # Subtraction node (prefilter output - feedback)
            u_unbounded[k] = prefilter_output - state_feedback
            
            # Nonlinear element (saturation block)
            u_bounded[k] = np.clip(u_unbounded[k], feed_min - ui_values["substrate_feed_initial"], 
                                   feed_max - ui_values["substrate_feed_initial"])
            #(Feedrate Antiwindup)
            substrate_feeding[k] = u_bounded[k] + ui_values["substrate_feed_initial"]
            
            # State Space Matrix B (applies to bounded control signal)
            if B.ndim == 2 and B.shape[1] == 1:  # B is a column vector
                Bu = B * u_bounded[k]  
            else:
                Bu = B @ np.array([[u_bounded[k]]])  # Matrix multiplication
            
            # State update equation (Matrix A + Matrix B input)
            # This combines the Matrix A block and the addition node before Delay 1/2
            x[k, :] = (A @ state_vector + Bu).flatten()
            
            # Matrix C (output equation)
            delta_output = float((C @ x[k, :].reshape(-1, 1))[0, 0])
            
            # Addition of initial gas flow rate to get actual gas flow (last addition node)
            biogas[k] = ui_values["gas_flow_initial"] + delta_output
            
            # Debug the first 3 iterations
            if k < 3:
                print(f"\n[STEP {k}]")
                print(f"  Prefilter output: {prefilter_output:.4f}")
                print(f"  State vector: {state_vector.flatten()}")
                print(f"  State feedback: {state_feedback:.4f}")
                print(f"  Control (unbounded): {u_unbounded[k]:.4f}")
                print(f"  Control (bounded): {u_bounded[k]:.4f}")
                print(f"  Delta output: {delta_output:.4f}")
                print(f"  Actual biogas output: {biogas[k]:.4f}")

        # Store simulation results
        global_vars['simulation_results'] = {
            'time': time,
            'biogas': biogas,
            'substrate_feeding': substrate_feeding,
            'u_unbounded': u_unbounded,
            'u_bounded': u_bounded,
            'x': x
        }

        print("\n[DEBUG] Simulation completed successfully. Plotting results...")

        # Plot results using anti-windup plotting function
        plot_with_antiwindup(global_vars['simulation_results'], ui_values)

        # Redraw canvas
        canvas.draw()

        # Display success message
        messagebox.showinfo("Simulation Complete", f"Simulation completed successfully for {days} days ({samples} steps).")

    except KeyError as e:
        messagebox.showerror("Key Error", f"Missing variable: {str(e)}")
        print(f"[ERROR] Key Error: {e}")

    except Exception as e:
        messagebox.showerror("Simulation Error", f"Error during simulation: {str(e)}")
        print(f"[ERROR] Simulation error: {e}")
        traceback.print_exc()


# Create simulation button
simulate_button = tk.Button(
    button_frame,
    text="Start Simulation", 
    font=("Arial", 10),
    bd=1, 
    relief="solid",
    command=run_simulation
)
simulate_button.pack(side='left', padx=10, pady=5)

# Add hover effects
def on_enter_sim(event):
    event.widget['background'] = '#d0eaff'  # Light blue color on hover

def on_leave_sim(event):
    event.widget['background'] = 'SystemButtonFace'  # Default button color

simulate_button.bind("<Enter>", on_enter_sim)
simulate_button.bind("<Leave>", on_leave_sim)

# Create a function to handle tab navigation
def navigate_to_feeding_schedule():
    """
    Navigate to the next tab (Feeding Schedule tab)
    """
    # Get the current tab index
    current_index = notebook.index(notebook.select())
    
    # Move to the next tab
    next_index = current_index + 1
    
    # Make sure the next index is within range
    if next_index < notebook.index("end"):
        notebook.select(next_index)
    else:
        print("No next tab available")

# Add a feeding schedule button to the button frame
feeding_schedule_button = tk.Button(
    button_frame,
    text="Feeding Schedule", 
    font=("Arial", 10),
    bd=1, 
    relief="solid",
    command=navigate_to_feeding_schedule
)
feeding_schedule_button.pack(side='right', padx=10, pady=5)

# Add hover effects for the new button
def on_enter_schedule(event):
    event.widget['background'] = '#d0ffea'  # Light green color on hover

def on_leave_schedule(event):
    event.widget['background'] = 'SystemButtonFace'  # Default button color

feeding_schedule_button.bind("<Enter>", on_enter_schedule)
feeding_schedule_button.bind("<Leave>", on_leave_schedule)

###################################################################################################
###################################################################################################
# FEEDING SCHEDULE

# Configure the Feeding Schedule tab
feeding_tab.grid_rowconfigure(0, weight=1)
feeding_tab.grid_columnconfigure(0, weight=1)
feeding_tab.grid_columnconfigure(1, weight=1)

# Create frames for daily and hourly feeding schedules
daily_frame = ttk.LabelFrame(feeding_tab, text="Daily feeding schedule", padding=10)
daily_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

hourly_frame = ttk.LabelFrame(feeding_tab, text="Two-hourly feeding schedule", padding=10)
hourly_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

# Create treeview for daily feeding schedule
daily_columns = ('Time', 'GPS [t]', 'Corn [t]', 'Manure [t]', 'Fodder residues [t]', 'Sum [t]')
daily_tree = ttk.Treeview(daily_frame, columns=daily_columns, show='headings', height=10)

# Configure daily treeview columns
for col in daily_columns:
    daily_tree.heading(col, text=col)
    daily_tree.column(col, width=100, anchor='center')

daily_tree.grid(row=0, column=0, sticky="nsew")
daily_scrollbar = ttk.Scrollbar(daily_frame, orient="vertical", command=daily_tree.yview)
daily_scrollbar.grid(row=0, column=1, sticky="ns")
daily_tree.configure(yscrollcommand=daily_scrollbar.set)

# Create treeview for hourly feeding schedule
hourly_columns = ('Time', 'GPS [t]', 'Corn [t]', 'Manure [t]', 'Fodder residues [t]', 'Sum [t]')
hourly_tree = ttk.Treeview(hourly_frame, columns=hourly_columns, show='headings', height=20)

# Configure hourly treeview columns
for col in hourly_columns:
    hourly_tree.heading(col, text=col)
    hourly_tree.column(col, width=100, anchor='center')

hourly_tree.grid(row=0, column=0, sticky="nsew")
hourly_scrollbar = ttk.Scrollbar(hourly_frame, orient="vertical", command=hourly_tree.yview)
hourly_scrollbar.grid(row=0, column=1, sticky="ns")
hourly_tree.configure(yscrollcommand=hourly_scrollbar.set)

# Configure daily and hourly frames for expansion
daily_frame.grid_rowconfigure(0, weight=1)
daily_frame.grid_columnconfigure(0, weight=1)
hourly_frame.grid_rowconfigure(0, weight=1)
hourly_frame.grid_columnconfigure(0, weight=1)

def feeding_plan_ui(start_date, simulation_results=None):
    """
    Calculate feeding plans based on simulation results - FIXED to match MATLAB output
    
    Parameters:
    - start_date: Start date for the feeding schedule
    - simulation_results: Results from the simulation (or None for default values)
    
    Returns:
    - Tuple of (None, sum_two_hourly, time_two_hourly, None, sum_daily, time_daily)
    """
    try:
        # Get the feeding proportions from the simulation
        # Define feeding proportions (from the original MATLAB code)
        anteil_GPS = 4/8     # 50% Grass Silage
        anteil_mais = 1.5/8  # 18.75% Corn
        anteil_mist = 2/8    # 25% Manure 
        anteil_reste = 0.5/8 # 6.25% Fodder residues
        
        # Create datetime objects for daily and two-hourly schedules
        if simulation_results is not None and 'time' in simulation_results:
            # If simulation results exist, use its timespan
            sim_duration_hours = simulation_results['time'][-1]
            days = int(sim_duration_hours / 24) + 1
        else:
            # No default values - abort if simulation results aren't available
            print("ERROR: No simulation time data available")
            return (None, np.array([]), [], None, np.array([]), [])
            
        print(f"Generating feeding schedule for {days} days starting from {start_date}")
            
       # Generate time points for daily schedule
        time_daily = [start_date + timedelta(days=i) for i in range(days)]
        time_two_hourly = []

        # Generate time points every 2 hours starting at 8:00 AM on the first day
        first_day_time = datetime.combine(
            start_date if not isinstance(start_date, datetime) else start_date.date(),
            time(hour=8, minute=0, second=0)
        )

        # Create the first time point at 8:00 AM
        time_two_hourly.append(first_day_time)

        # Generate the rest of the time points every 2 hours
        current_time = first_day_time
        for _ in range(days * 12 - 1):  # -1 because we already added the first time point
            current_time = current_time + timedelta(hours=2)
            time_two_hourly.append(current_time)
                
        # Retrieve substrate values - NEVER use default values
        if simulation_results is not None and 'substrate_feeding' in simulation_results:
            # Use ONLY actual simulation data - no hardcoded scaling factors
            # The correct values should come from the simulation directly
            
            substrate_daily = []
            
            # Aggregate 2-hour samples into daily totals
            for i in range(days):
                idx_start = i * 12  # 12 samples per day for 2-hour intervals
                idx_end = idx_start + 12
                
                if idx_start < len(simulation_results['substrate_feeding']):
                    # Sum the values for this day
                    daily_values = simulation_results['substrate_feeding'][idx_start:min(idx_end, len(simulation_results['substrate_feeding']))]
                    daily_sum = sum(daily_values)  # No scaling - use actual values
                    substrate_daily.append(daily_sum)
                else:
                    # If we're beyond simulation data, use last day's value
                    print(f"ERROR: Missing simulation data for day {i+1}")
                    print(f"Required index range {idx_start}-{idx_end}, but simulation data has only {len(simulation_results['substrate_feeding'])} entries")
                    # Return partial data
                    return (None, np.array([]), [], None, np.array(substrate_daily), time_daily[:len(substrate_daily)])
            
            sum_daily = np.array(substrate_daily)
            
            # Create two-hourly values directly from simulation data
            sum_two_hourly = []
            
            # Process each day
            for day_idx in range(days):
                day_start = day_idx * 12  # 12 samples per day for 2-hour intervals
                
                if day_start < len(simulation_results['substrate_feeding']):
                    # Get actual hourly distribution from simulation data
                    day_values = simulation_results['substrate_feeding'][day_start:min(day_start+12, len(simulation_results['substrate_feeding']))]
                    
                    # Add the values directly - simulation already has 2-hour intervals
                    two_hour_values = day_values.copy()
                    
                    # If we don't have all 12 values for the day, fill the remaining with the last value
                    if len(two_hour_values) < 12:
                        print(f"WARNING: Only {len(two_hour_values)} data points for day {day_idx+1}, filling remaining with last value")
                        last_value = two_hour_values[-1] if two_hour_values else 0
                        two_hour_values.extend([last_value] * (12 - len(two_hour_values)))
                    
                    # Add values to the array
                    sum_two_hourly.extend(two_hour_values)
                else:
                    # Error if insufficient data - no defaults
                    print(f"ERROR: Missing simulation data for day {day_idx+1}")
                    return (None, np.array(sum_two_hourly), time_two_hourly[:len(sum_two_hourly)], 
                            None, sum_daily[:day_idx], time_daily[:day_idx])
                    
        else:
            # Instead of using default values, print an error message
            print("ERROR: No simulation results available for feeding schedule")
            print("Make sure to run simulation first and verify substrate_feeding data exists")
            # Return empty arrays with error indication
            return (None, np.array([]), [], None, np.array([]), [])
        
        # Check array lengths and print errors instead of adjusting with defaults
        if len(sum_daily) < days:
            print(f"ERROR: Insufficient daily values. Expected {days}, got {len(sum_daily)}")
            # Adjust days to match available data
            days = len(sum_daily)
            # Truncate time arrays to match
            time_daily = time_daily[:days]
            time_two_hourly = time_two_hourly[:days*12]
            
        elif len(sum_daily) > days:
            print(f"WARNING: Too many daily values. Expected {days}, got {len(sum_daily)}")
            # Truncate to requested days
            sum_daily = sum_daily[:days]
            
        if len(sum_two_hourly) < days * 12:
            print(f"ERROR: Insufficient two-hourly values. Expected {days*12}, got {len(sum_two_hourly)}")
            # Truncate time array to match available data
            time_two_hourly = time_two_hourly[:len(sum_two_hourly)]
            
        elif len(sum_two_hourly) > days * 12:
            print(f"WARNING: Too many two-hourly values. Expected {days*12}, got {len(sum_two_hourly)}")
            # Truncate to requested number
            sum_two_hourly = sum_two_hourly[:days*12]
            
        # Print summary for debugging
        print(f"Generated {len(time_daily)} daily values and {len(time_two_hourly)} two-hourly values")
        print(f"Daily sums: {sum_daily}")
        print(f"Two-hourly sample: {sum_two_hourly[:12]}")
            
        return (None, sum_two_hourly, time_two_hourly, None, sum_daily, time_daily)
        
    except Exception as e:
        print(f"Error in feeding_plan_ui: {e}")
        traceback.print_exc()
        return (None, np.array([]), [], None, np.array([]), [])

def update_feeding_schedule():
    """
    Update the feeding schedule tables based on the current simulation results - FIXED
    """
    try:
        # First check if simulation results exist
        if 'simulation_results' not in global_vars:
            print("ERROR: No simulation results available")
            print("Please run the simulation before generating the feeding schedule")
            messagebox.showerror("Error", "No simulation results available. Please run the simulation first.")
            return
            
        # Check if substrate_feeding exists in simulation results
        if 'substrate_feeding' not in global_vars['simulation_results']:
            print("ERROR: No substrate feeding data in simulation results")
            print("simulation_results keys:", global_vars['simulation_results'].keys())
            messagebox.showerror("Error", "No substrate feeding data in simulation results. Check simulation configuration.")
            return
            
        # Clear existing data in the trees
        for item in daily_tree.get_children():
            daily_tree.delete(item)
            
        for item in hourly_tree.get_children():
            hourly_tree.delete(item)
        
        # Get start date from control tab
        start_date = None
        
        # Get UI values first
        ui_values = get_ui_values()
        
        # Get the start date with clear error handling
        if hasattr(start_date_field, 'get_date'):
            start_date = start_date_field.get_date()
            print(f"Using start date from date field: {start_date}")
        elif 'start_date' in ui_values:
            start_date = ui_values['start_date']
            print(f"Using start date from UI values: {start_date}")
        else:
            print("ERROR: Could not determine start date for feeding schedule")
            print("Please ensure start_date_field is properly configured or simulation has a start_date")
            messagebox.showerror("Error", "Could not determine start date. Check your settings.")
            return
        
        # Get simulation results if available
        simulation_results = global_vars.get('simulation_results', None)
        
        # Print debug info about simulation results
        print(f"Simulation results summary:")
        print(f"- Time points: {len(global_vars['simulation_results']['time'])}")
        print(f"- Substrate feeding points: {len(global_vars['simulation_results']['substrate_feeding'])}")
        print(f"- First few substrate values: {global_vars['simulation_results']['substrate_feeding'][:5]}")
        
        # Calculate feeding plan
        result = feeding_plan_ui(start_date, global_vars['simulation_results'])
        _, sum_two_hourly, time_two_hourly, _, sum_daily, time_daily = result
        
        # Check if data was returned
        if len(sum_daily) == 0 or len(sum_two_hourly) == 0:
            print("ERROR: No feeding schedule data was generated")
            messagebox.showerror("Error", "Failed to generate feeding schedule data. See debug output for details.")
            return
            
        # Define feeding proportions (use exact values from MATLAB code)
        anteil_GPS = 4/8     # 50% Grass Silage
        anteil_mais = 1.5/8  # 18.75% Corn
        anteil_mist = 2/8    # 25% Manure 
        anteil_reste = 0.5/8 # 6.25% Fodder residues
        
        print(f"Feeding proportions: GPS={anteil_GPS}, Corn={anteil_mais}, Manure={anteil_mist}, Residue={anteil_reste}")
        
        # Update daily feeding schedule
        print("\nDaily feeding schedule data:")
        for i, (time, sum_val) in enumerate(zip(time_daily, sum_daily)):
            # Format datetime to include both date and time: "15-Nov-2024 06:00"
            time_str = time.strftime("%d-%b-%Y %H:%M")
            
            # Calculate individual feedstock values
            gps_val = anteil_GPS * sum_val
            corn_val = anteil_mais * sum_val
            manure_val = anteil_mist * sum_val
            residue_val = anteil_reste * sum_val
            
            # Debug print
            print(f"Day {i+1}: {time_str}, Total: {sum_val:.4f}, GPS: {gps_val:.4f}, Corn: {corn_val:.4f}, "
                  f"Manure: {manure_val:.4f}, Residue: {residue_val:.4f}")
            
            # Insert into treeview with 4 decimal precision to match MATLAB
            daily_tree.insert('', tk.END, values=(
                time_str,
                f"{gps_val:.4f}",
                f"{corn_val:.4f}",
                f"{manure_val:.4f}",
                f"{residue_val:.4f}",
                f"{sum_val:.4f}"
            ))
        
        # Update two-hourly feeding schedule
        print("\nFirst few entries of two-hourly feeding schedule:")
        max_debug_entries = min(12, len(time_two_hourly))  # Print first 12 entries or fewer
        
        for i, (time, sum_val) in enumerate(zip(time_two_hourly, sum_two_hourly)):
            # Format datetime to include both date and time: "15-Nov-2024 06:00:00"
            time_str = time.strftime("%d-%b-%Y %H:%M:%S")
            
            # Calculate individual feedstock values
            gps_val = anteil_GPS * sum_val
            corn_val = anteil_mais * sum_val
            manure_val = anteil_mist * sum_val
            residue_val = anteil_reste * sum_val
            
            # Debug print for first few entries
            if i < max_debug_entries:
                print(f"Hour {i+1}: {time_str}, Total: {sum_val:.4f}, GPS: {gps_val:.4f}, "
                      f"Corn: {corn_val:.4f}, Manure: {manure_val:.4f}, Residue: {residue_val:.4f}")
            
            # Insert into treeview with 4 decimal precision to match MATLAB
            hourly_tree.insert('', tk.END, values=(
                time_str,
                f"{gps_val:.4f}",
                f"{corn_val:.4f}",
                f"{manure_val:.4f}",
                f"{residue_val:.4f}",
                f"{sum_val:.4f}"
            ))
            
    except Exception as e:
        print(f"Error updating feeding schedule: {e}")
        traceback.print_exc()

# Create a frame for buttons
button_frame = ttk.Frame(feeding_tab, padding=10)
button_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

# Add a button to update the feeding schedule
update_button = tk.Button(
    button_frame,
    text="Update Feeding Schedule",
    font=("Arial", 10),
    bd=1,
    relief="solid",
    command=update_feeding_schedule
)
update_button.pack(side='left', padx=10, pady=5)

# Create a button to export the feeding schedule to Excel
def export_to_excel():
    """Export the feeding schedule to an Excel file"""
    try:
        # Get file name from user
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Feeding Schedule As"
        )
        
        if not file_path:
            return  # User cancelled
            
        # Create a new workbook
        wb = load_workbook() if 'load_workbook' in globals() else None
        
        if wb is None:
            messagebox.showerror("Export Error", "Excel export functionality not available.")
            return
            
        # Create sheets for daily and hourly schedules
        daily_sheet = wb.create_sheet("Daily Feeding Schedule")
        hourly_sheet = wb.create_sheet("Two-Hourly Feeding Schedule")
        
        # Add headers
        headers = ['Time', 'GPS [t]', 'Corn [t]', 'Manure [t]', 'Fodder residues [t]', 'Sum [t]']
        daily_sheet.append(headers)
        hourly_sheet.append(headers)
        
        # Add daily data
        for item_id in daily_tree.get_children():
            values = daily_tree.item(item_id, 'values')
            daily_sheet.append(values)
            
        # Add hourly data
        for item_id in hourly_tree.get_children():
            values = hourly_tree.item(item_id, 'values')
            hourly_sheet.append(values)
            
        # Save the workbook
        wb.save(file_path)
        messagebox.showinfo("Export Successful", f"Feeding schedule exported to {file_path}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Error exporting to Excel: {str(e)}")
        print(f"Error exporting to Excel: {e}")
        traceback.print_exc()

# Add Export button
export_button = tk.Button(
    button_frame,
    text="Export to Excel",
    font=("Arial", 10),
    bd=1,
    relief="solid",
    command=export_to_excel
)
export_button.pack(side='left', padx=10, pady=5)

# Add hover effects
update_button.bind("<Enter>", on_enter_sim)
update_button.bind("<Leave>", on_leave_sim)
export_button.bind("<Enter>", on_enter_sim)
export_button.bind("<Leave>", on_leave_sim)

# Update the feeding schedule when switching to the tab
def on_tab_selected(event):
    """Update the feeding schedule when the tab is selected"""
    selected_tab = event.widget.select()
    tab_text = event.widget.tab(selected_tab, "text")
    
    if tab_text == "Feeding Schedule":
        update_feeding_schedule()

# Bind the tab selection event
notebook.bind("<<NotebookTabChanged>>", on_tab_selected)

# Initial update of the feeding schedule
update_feeding_schedule()

# Run Tkinter Main Loop
root.mainloop() 




